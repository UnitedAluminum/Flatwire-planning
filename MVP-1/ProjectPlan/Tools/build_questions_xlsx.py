"""
Build the client-facing questions workbook.

    python MVP-1/ProjectPlan/Tools/build_questions_xlsx.py [output.xlsx]

With no arguments it writes MVP-1/SRS/FlatWire_ClientQuestions.xlsx, which is the default below.

WHY THIS EXISTS
    United Aluminum owes answers on the open questions and confirmation of the recorded
    decisions. Both registers under Analysis/ are written for the build team and are dense
    with table names, constraint names, endpoints and requirement identifiers - none of which
    may appear in a client deliverable. The client-facing rewrite lives in
    MVP-1/RequirementDocuments/ClientQuestionsContent.md and this script merges it with the
    registers. Edit the markdown and re-run - never edit the .xlsx.

WHERE EACH FIELD COMES FROM
    Structure  - question number, priority, scope, owner, decided date - is parsed from the
                 index tables of the two registers, so it cannot drift from them.
    Prose      - the plain-language question, background, options, recommendation - comes from
                 the content file, which is the only place it is authored.
    Nothing is duplicated between the two. "Register title" is read from the content file for
    the drift guard alone and is never written to the workbook.

FOUR GUARDS, ALL FATAL
    1. Coverage      - every question in a register has a content entry and vice versa.
    2. Drift         - the content entry's recorded register title still matches the register.
                       Catches a renumbering silently moving prose onto the wrong question.
    3. Team names    - "Needs input from" carries no Nagarro-side name.
    4. Leakage       - every cell of the built workbook is scanned for file names, paths,
                       requirement/gap/test identifiers, table names, endpoints, code spans
                       and machine tag paths. The tag surface has exactly one client-facing
                       home and this workbook is not it (see the anti-drift rule in CLAUDE.md).

REQUIREMENTS
    openpyxl (tested against 3.1.5). Nothing else.

GOTCHA
    If the output is open in Excel the write fails with PermissionError - a ~$ lock file in
    MVP-1/SRS/ is the tell. Close it and re-run.
"""
import os, re, sys

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, '..', '..', '..'))

OPEN_REGISTER = os.path.join(ROOT, 'Analysis', 'FlatWireOpenQuestions.md')
DECIDED_REGISTER = os.path.join(ROOT, 'Analysis', 'FlatWireDecidedQuestions.md')
CONTENT = os.path.join(ROOT, 'MVP-1', 'RequirementDocuments', 'ClientQuestionsContent.md')
DEFAULT_OUT = os.path.join(ROOT, 'MVP-1', 'SRS', 'FlatWire_ClientQuestions.xlsx')

ISSUE_DATE = 'August 12, 2026'
RETURN_BY = 'To be agreed'          # deliberately not invented - see the plan's out-of-scope note
RETURN_TO = 'To be confirmed'

# Nagarro-side names. "Needs input from" names the client stakeholders only.
TEAM_NAMES = ('Jaspreet', 'Srikanth', 'Shray', 'Yogender')

RESPONSE_OPEN = '"Agree with recommendation,Different answer - see next column,Needs discussion"'
RESPONSE_DECIDED = '"Confirmed,Not confirmed - see next column,Needs discussion"'

# ----------------------------------------------------------------- markdown helpers

def clean(text):
    """Strip the markdown a register table cell carries."""
    text = re.sub(r'\[([^\]]*)\]\([^)]*\)', r'\1', text)   # links -> label
    text = text.replace('`', '').replace('**', '').replace('*', '')
    return re.sub(r'\s+', ' ', text).strip()


_BOLD = InlineFont(rFont='Calibri', sz=11, b=True)
_ITALIC = InlineFont(rFont='Calibri', sz=11, i=True)
_EMPHASIS = re.compile(r'\*\*(.+?)\*\*|(?<![\w*])\*([^*\n]+?)\*(?![\w*])')


def _fold_blank_runs(parts):
    """Never emit a whitespace-only run.

    openpyxl 3.1.5 writes xml:space="preserve" on a whitespace-EDGED run (' x ') but
    omits it on a whitespace-ONLY one (' '), so XML normalisation strips the content and
    Excel reports 'Repaired Records: String properties' on open. Two **bold** spans
    separated by a single space produce exactly that run.

    The whitespace is folded into an adjacent run instead of being dropped, so the text
    is unchanged. Folding it into an emphasised run is harmless - a bold space is
    visually identical to a plain one.
    """
    def attach(part, ws, before=False):
        if isinstance(part, str):
            return (ws + part) if before else (part + ws)
        text = (ws + part.text) if before else (part.text + ws)
        return TextBlock(part.font, text)

    out, pending = [], ''
    for part in parts:
        if isinstance(part, str):
            part = pending + part
            pending = ''
            if part and not part.strip():
                if out:
                    out[-1] = attach(out[-1], part)
                else:
                    pending = part          # nothing to fold back into yet
                continue
            if out and isinstance(out[-1], str):
                out[-1] += part
            else:
                out.append(part)
        else:
            if pending:
                part, pending = attach(part, pending, before=True), ''
            out.append(part)
    if pending:
        out[-1] = attach(out[-1], pending) if out else pending
    return out


def render(text):
    """Turn authored markdown emphasis into Excel rich text.

    The prose is authored with **bold** carrying real weight - it is what makes a
    400-character cell scannable - so it is rendered rather than stripped. Backticks are
    dropped: a code span has no meaning in a client deliverable, and the leakage scan
    rejects them anyway.
    """
    text = text.replace('`', '')
    if not _EMPHASIS.search(text):
        return text
    parts, pos = [], 0
    for m in _EMPHASIS.finditer(text):
        if m.start() > pos:
            parts.append(text[pos:m.start()])
        if m.group(1) is not None:
            parts.append(TextBlock(_BOLD, m.group(1)))
        else:
            parts.append(TextBlock(_ITALIC, m.group(2)))
        pos = m.end()
    if pos < len(text):
        parts.append(text[pos:])
    return CellRichText(*_fold_blank_runs(parts))


def norm(text):
    """Normalise a title for the drift guard: letters, digits and single spaces only."""
    text = clean(text).lower()
    text = re.sub(r'[^a-z0-9 ]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def parse_index_table(path, first_header):
    """Pull an index table keyed on a leading question number out of a register."""
    rows, in_table, header = {}, False, None
    with open(path, encoding='utf-8') as fh:
        for line in fh:
            line = line.rstrip('\n')
            if not in_table:
                # Must be the index table proper: a leading '#' column plus the named header.
                # Both registers carry other tables containing these words.
                if re.match(r'^\|\s*#\s*\|', line) and first_header in line:
                    header = [clean(c) for c in line.strip('|').split('|')]
                    in_table = True
                continue
            if not line.startswith('|'):
                break
            if re.match(r'^\|[\s:|-]+\|$', line):
                continue
            cells = [clean(c) for c in line.strip('|').split('|')]
            if not cells or not cells[0].isdigit():
                break
            rows[int(cells[0])] = dict(zip(header, cells))
    if not rows:
        sys.exit(f'FATAL: no index table found in {os.path.relpath(path, ROOT)}')
    return rows


FIELDS_OPEN = ('Register title', 'Area', 'Needs input from', 'What we need',
               'Answer together with', 'Question', 'Background', 'Already agreed',
               'Options', 'Recommended answer', 'Why', 'Impact if unanswered')
FIELDS_DECIDED = ('Register title', 'Area', 'Question', 'Background',
                  'Decision as recorded', 'Still open', 'Our recommendation')
# "Recommended answer" is optional: a question may deliberately go to the client with no
# vendor answer attached. Q10 (footage-to-weight) is the first such case - see the note
# above the filtered index in FlatWireOpenQuestions.md. "Why" is optional for the same
# reason and only that reason: it exists to justify the recommendation, so a question
# carrying no recommendation has nothing for it to say (Q4, skid labelling).
REQUIRED_OPEN = tuple(f for f in FIELDS_OPEN
                      if f not in ('Answer together with', 'Already agreed', 'Options',
                                   'Recommended answer', 'Why'))
REQUIRED_DECIDED = tuple(f for f in FIELDS_DECIDED if f != 'Still open')


def parse_content(path):
    """Parse the authored content file into {part: {qnum: {field: text}}}."""
    text = open(path, encoding='utf-8').read()
    parts, current = {'open': {}, 'decided': {}}, None
    known = set(FIELDS_OPEN) | set(FIELDS_DECIDED)

    # Split on the two part headers, then on '## Q##' inside each.
    for chunk in re.split(r'^# Part \d+ — (.+)$', text, flags=re.M)[1:]:
        if chunk.strip().startswith('Open Questions'):
            current = 'open'; continue
        if chunk.strip().startswith('Decisions to Confirm'):
            current = 'decided'; continue
        if current is None:
            continue
        blocks = re.split(r'^## Q(\d+)\s*$', chunk, flags=re.M)
        for qnum, body in zip(blocks[1::2], blocks[2::2]):
            entry, field = {}, None
            for line in body.split('\n'):
                if line.strip() in ('---', ''):
                    if line.strip() == '---':
                        field = None
                    elif field:
                        entry[field] += '\n'
                    continue
                m = re.match(r'^\*\*([A-Z][^:*]*):\*\*\s?(.*)$', line)
                if m and m.group(1) in known:
                    field = m.group(1)
                    entry[field] = m.group(2).strip()
                elif field:
                    entry[field] += ' ' + line.strip()
            unknown = set(entry) - known
            if unknown:
                sys.exit(f'FATAL: Q{qnum} has unknown field(s): {sorted(unknown)}')
            parts[current][int(qnum)] = {k: re.sub(r'\s+\n', '\n', v).strip()
                                         for k, v in entry.items()}
    if not parts['open'] or not parts['decided']:
        sys.exit('FATAL: content file is missing one of its two parts')
    return parts


# ----------------------------------------------------------------------- guards

def guard(open_reg, decided_reg, content):
    errors = []

    for part, reg, required, title_key in (
            ('open', open_reg, REQUIRED_OPEN, 'Question (Short)'),
            ('decided', decided_reg, REQUIRED_DECIDED, 'Subject')):
        missing = sorted(set(reg) - set(content[part]))
        extra = sorted(set(content[part]) - set(reg))
        for q in missing:
            errors.append(f'Q{q} is in the {part} register with no content entry')
        for q in extra:
            errors.append(f'Q{q} has a content entry but is not in the {part} register')

        for q, entry in sorted(content[part].items()):
            for f in required:
                if not entry.get(f):
                    errors.append(f'Q{q} is missing required field "{f}"')
            if q not in reg:
                continue
            a, b = norm(entry.get('Register title', '')), norm(reg[q][title_key])
            n = min(20, len(a), len(b))
            if n == 0 or a[:n] != b[:n]:
                errors.append(
                    f'Q{q} drift: content says "{clean(entry.get("Register title", ""))}" '
                    f'but the register says "{clean(reg[q][title_key])}" '
                    f'- a renumbering would put prose on the wrong question')
            who = entry.get('Needs input from', '')
            for name in TEAM_NAMES:
                if name in who:
                    errors.append(f'Q{q} "Needs input from" names our own team: {name}')

    if errors:
        print('FATAL - build refused:')
        for e in errors:
            print('  *', e)
        sys.exit(1)


LEAKS = [
    (r'\.(?:md|html|sql|docx|xlsx|py|js|scss|css)\b', 'file name'),
    (r'\b(?:MVP-1|MVP-2|Analysis|BaseDocuments|LatestDocument|DevelopmentPlan|'
     r'RequirementDocuments|Mockups|DBChanges|ProjectPlan|ShopfloorPlan)/', 'repository path'),
    (r'\bFR-\d', 'requirement identifier'),
    (r'\bOI-\d', 'open-item identifier'),
    (r'\bG\d{1,2}\b', 'gap identifier'),
    (r'\bTC-\d', 'test-case identifier'),
    (r'\bFW-\d', 'backlog identifier'),
    (r'\bPLC-Q', 'PLC question identifier'),
    (r'\bPSG-', 'generation-spec identifier'),
    (r'\bD-\d{2}\b', 'decision identifier'),
    (r'\b(?:PCI|CHK|WLD|TRV|PSM|FRT|DM|SQ)\d{2,3}\b', 'requirement identifier'),
    (r'\b(?:CK|IX|UX|FK)_\w', 'constraint or index name'),
    (r'\b(?:POST|GET|PUT|PATCH|DELETE)\s+/', 'endpoint'),
    (r'`', 'code span'),
    (r'\bFL[123]\.\w', 'machine tag path'),
    (r'\bDashboard\s*\d', 'screen number'),
    (r'\b(?:RodStaging|AlloyProperty|CoilTraceability|CoilOutput|FlatWireRun|PassSchedule'
     r'|SpoolCheckin|RodCheckin|RodCheckout|WipRejection|SpcCheckpoint|SpcMeasurement'
     r'|RunReading|WeldEvent|RollOverride|DieChangeEvent|RunPauseEvent|PayoffPosition'
     r'|FlatwireQueue|planning_routings|wip_coil_orders|FlatWireDB|united_db|CommonDB)\b',
     'table or column name'),
]


def scan_for_leaks(path):
    wb = load_workbook(path)
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                for pattern, label in LEAKS:
                    m = re.search(pattern, cell.value)
                    if m:
                        hits.append((ws.title, cell.coordinate, label, m.group(0),
                                     cell.value[max(0, m.start() - 40):m.end() + 40]))
    return hits


# ------------------------------------------------------------------- formatting

HEAD_FILL = PatternFill('solid', fgColor='1F3864')
TITLE_FONT = Font(name='Calibri', size=15, bold=True, color='1F3864')
SUB_FONT = Font(name='Calibri', size=10, italic=True, color='404040')
HEAD_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
BODY_FONT = Font(name='Calibri', size=11)
RESP_FILL = PatternFill('solid', fgColor='FFF7E0')
ALT_FILL = PatternFill('solid', fgColor='F5F7FA')
PRIORITY_FILL = {
    'Critical': PatternFill('solid', fgColor='F8CBCB'),
    'High': PatternFill('solid', fgColor='FBE3C0'),
    'Medium': PatternFill('solid', fgColor='D9E7F5'),
    'Low': PatternFill('solid', fgColor='E4E7EA'),
}
THIN = Side(style='thin', color='C8CDD3')
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
HEAD_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)


def write_sheet(ws, title, subtitle, columns, rows, response_from, validation_formula):
    """columns: list of (header, width). response_from: 0-based index of the first client column."""
    ws.sheet_view.showGridLines = False
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 22
    ws['A2'] = subtitle
    ws['A2'].font = SUB_FONT
    ws.row_dimensions[2].height = 14

    header_row = 4
    for i, (head, width) in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=i, value=head)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = HEAD_ALIGN
        c.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[header_row].height = 30

    for r, values in enumerate(rows, start=header_row + 1):
        for i, value in enumerate(values, start=1):
            c = ws.cell(row=r, column=i, value=render(value) if value else None)
            c.font = BODY_FONT
            c.alignment = TOP_WRAP
            c.border = CELL_BORDER
            if i - 1 >= response_from:
                c.fill = RESP_FILL
            elif (r - header_row) % 2 == 0:
                c.fill = ALT_FILL
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='top')

    last_row = header_row + len(rows)
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(columns))}{last_row}'
    ws.freeze_panes = f'B{header_row + 1}'

    dv = DataValidation(type='list', formula1=validation_formula, allow_blank=True,
                        showDropDown=False, errorStyle='warning')
    dv.error = 'Pick from the list, or type your own note in the next column.'
    dv.prompt = 'Choose one.'
    ws.add_data_validation(dv)
    col = get_column_letter(response_from + 1)
    dv.add(f'{col}{header_row + 1}:{col}{last_row}')
    return last_row


def priority_column(ws, rows, header_row, col_index):
    for r, values in enumerate(rows, start=header_row + 1):
        value = values[col_index]
        if value in PRIORITY_FILL:
            ws.cell(row=r, column=col_index + 1).fill = PRIORITY_FILL[value]


# ------------------------------------------------------------------------ build

def build(out_path):
    open_reg = parse_index_table(OPEN_REGISTER, 'Question (Short)')
    decided_reg = parse_index_table(DECIDED_REGISTER, 'Subject')
    content = parse_content(CONTENT)
    guard(open_reg, decided_reg, content)

    wb = Workbook()

    # ---------------------------------------------------------------- Read Me
    ws = wb.active
    ws.title = 'Read Me'
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '1F3864'
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 108

    n_open, n_decided = len(content['open']), len(content['decided'])
    priorities = [clean(open_reg[q]['Priority']) for q in sorted(open_reg)]
    counts = ' · '.join(f'{p} {priorities.count(p)}'
                        for p in ('Critical', 'High', 'Medium', 'Low')
                        if priorities.count(p))

    readme = [
        ('TITLE', 'Flat Wire Mill — Questions for United Aluminum'),
        ('SUB', f'Issued {ISSUE_DATE} · {n_open} open questions · '
                f'{n_decided} recorded decisions to confirm'),
        ('GAP', ''),
        ('H', 'What we are asking for'),
        ('P', 'Two things, on the two sheets that follow. On "Decisions to Confirm" we have '
              'written back the decisions already taken in our meetings and calls, and we are '
              'asking you to confirm each one still stands. On "Open Questions" we have set out '
              'what is still outstanding, each with a recommended answer, and we are asking you '
              'to agree it or give us the correct answer.'),
        ('P', 'Every question carries a recommendation. That is our proposed answer — it is not '
              'a decision and it carries no authority. It exists so this is a confirm-or-correct '
              'exercise rather than a blank page, and so that we have a defensible default to '
              'build to while an answer is outstanding. Confirming a recommendation is what '
              'closes the question.'),
        ('GAP', ''),
        ('H', 'How to fill it in'),
        ('P', 'Use the shaded columns on the right of each sheet. Pick a value from the dropdown, '
              'and where you disagree or want to qualify the answer, put the correct answer in the '
              'column beside it — free text, as long as you like. Add your name and the date so we '
              'know who to go back to. Everything to the left of the shaded columns is ours; '
              'please leave it as it is so we can see what you were answering.'),
        ('P', 'Both sheets have filters on the header row. Filter by priority to work the urgent '
              'items first, or by "Needs input from" to route questions to the right person.'),
        ('GAP', ''),
        ('H', 'Answer together with'),
        ('P', 'Some questions are one conversation rather than several. Where that is the case the '
              '"Answer together with" column lists the others, so they can be taken in one pass. '
              'The machine-interface questions in particular are all blocked on the same '
              'controls engineer and are best closed in a single session.'),
        ('GAP', ''),
        ('H', 'What the priorities mean'),
        ('K', 'Critical — needed before development starts on the part of the system that depends on it.'),
        ('K', 'High — needed before user acceptance testing, late September 2026.'),
        ('K', 'Medium — needed before production go-live, fourth quarter 2026.'),
        ('K', 'Low — can be settled after go-live.'),
        ('P', f'Open questions by priority: {counts}.'),
        ('GAP', ''),
        ('H', 'What the columns mean'),
        ('K', '"What we need" — whether we are asking for a decision, for figures or data we do '
              'not have, or for confirmation that a reading of ours is correct.'),
        ('K', '"Already agreed" — the part of a question that is already settled. Please do not '
              're-answer it; the Question column asks only what is still outstanding.'),
        ('K', '"Anything still open" on the decisions sheet — a decision that is closed in '
              'substance but left a residual we still need. Several do.'),
        ('K', '"Impact if unanswered" — what the question is holding up, so the sequence of '
              'answers can be prioritised.'),
        ('GAP', ''),
        ('H', 'Two notes on reading it'),
        ('K', 'A handful of decisions apply only in part to the current release, and each says so '
              'in its own row. Please read and confirm the half that applies.'),
        ('K', 'One open question — the finishing-mill dancer modes — carries a recommendation we '
              'believe may be wrong, because it argues against something you told us earlier. It '
              'is flagged in the row and we would rather show you the conflict than resolve it '
              'ourselves.'),
        ('GAP', ''),
        ('H', 'Returning it'),
        ('K', f'Return to: {RETURN_TO}'),
        ('K', f'Return by: {RETURN_BY}'),
    ]

    r = 1
    for kind, text in readme:
        if kind == 'GAP':
            r += 1
            continue
        cell = ws.cell(row=r, column=1 if kind in ('TITLE', 'SUB', 'H') else 2)
        if kind == 'TITLE':
            cell.value = text; cell.font = TITLE_FONT; ws.row_dimensions[r].height = 22
        elif kind == 'SUB':
            cell.value = text; cell.font = SUB_FONT
        elif kind == 'H':
            cell.value = text
            cell.font = Font(name='Calibri', size=12, bold=True, color='1F3864')
        else:
            cell.value = render(text)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            ws.row_dimensions[r].height = None
        r += 1

    # -------------------------------------------------- Decisions to Confirm
    ws = wb.create_sheet('Decisions to Confirm')
    ws.sheet_properties.tabColor = '2E7D32'
    cols = [('Q', 6), ('Area', 22), ('Question', 46), ('Background', 46),
            ('Decision as recorded', 76), ('Date recorded', 13),
            ('Anything still open', 50), ('Our recommendation', 44),
            ('Confirmed?', 20), ('If not confirmed, the correct answer', 34),
            ('Confirmed by', 16), ('Date', 12)]
    rows = []
    for q in sorted(content['decided']):
        e, reg = content['decided'][q], decided_reg[q]
        rows.append([f'Q{q}', e['Area'], e['Question'], e['Background'],
                     e['Decision as recorded'], reg.get('Decided', ''),
                     e.get('Still open', ''), e['Our recommendation'],
                     '', '', '', ''])
    write_sheet(ws, 'Decisions to Confirm',
                f'{len(rows)} decisions already taken, written back for your confirmation. '
                f'Please complete the shaded columns.',
                cols, rows, response_from=8, validation_formula=RESPONSE_DECIDED)

    # ------------------------------------------------------- Open Questions
    ws = wb.create_sheet('Open Questions')
    ws.sheet_properties.tabColor = 'B71C1C'
    cols = [('Q', 6), ('Area', 22), ('Priority', 10), ('Needs input from', 20),
            ('What we need', 16), ('Question', 52), ('Background', 60),
            ('Already agreed', 50), ('Options', 44), ('Recommended answer', 60),
            ('Why', 46), ('Impact if unanswered', 34), ('Answer together with', 18),
            ('Your response', 22), ('Your answer or correction', 34),
            ('Answered by', 16), ('Date', 12)]
    order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3}
    rows = []
    for q in sorted(content['open'],
                    key=lambda q: (order.get(clean(open_reg[q]['Priority']), 9), q)):
        e, reg = content['open'][q], open_reg[q]
        rows.append([f'Q{q}', e['Area'], clean(reg['Priority']), e['Needs input from'],
                     e['What we need'], e['Question'], e['Background'],
                     e.get('Already agreed', ''), e.get('Options', ''),
                     e.get('Recommended answer', ''), e.get('Why', ''),
                     e['Impact if unanswered'],
                     e.get('Answer together with', ''), '', '', '', ''])
    write_sheet(ws, 'Open Questions',
                f'{len(rows)} questions outstanding, almost all with our recommended '
                f'answer. Please complete the shaded columns. Sorted by priority.',
                cols, rows, response_from=13, validation_formula=RESPONSE_OPEN)
    priority_column(ws, rows, header_row=4, col_index=2)

    wb.save(out_path)
    return len(content['decided']), len(content['open'])


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT
    n_decided, n_open = build(out)
    rel = os.path.relpath(out, ROOT)
    print(f'wrote {rel}')
    print(f'  Decisions to Confirm : {n_decided} rows')
    print(f'  Open Questions       : {n_open} rows')

    hits = scan_for_leaks(out)
    if hits:
        print(f'\nFATAL - {len(hits)} leak(s) found in the built workbook:')
        for sheet, coord, label, found, context in hits[:40]:
            print(f'  {sheet}!{coord}  {label}: {found!r}')
            print(f'      ...{context}...')
        os.remove(out)
        sys.exit(1)
    print('  leakage scan         : clean')


if __name__ == '__main__':
    main()
