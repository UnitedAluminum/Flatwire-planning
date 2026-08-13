"""
Build the client-facing development plan workbook.

    python MVP-1/ProjectPlan/Tools/build_development_plan_xlsx.py [output.xlsx]

With no arguments it writes MVP-1/SRS/FlatWire_DevelopmentPlan.xlsx, which is the default below.

WHY THIS EXISTS
    The development plan is complete and reconciled, but every document holding it is written
    for the build team: keyed on internal identifiers, citing file paths and section numbers,
    naming components, endpoints and database tables. 68 of the 116 work-item titles trip the
    leakage guard below. A client sheet therefore cannot be produced by filtering the plan -
    the prose has to be written about the business outcome instead. That prose lives in
    MVP-1/ProjectPlan/Tools/DevelopmentPlanContent.md and this script merges it with the plan.
    Edit the markdown and re-run - never edit the .xlsx.

WHERE EACH FIELD COMES FROM
    Structure  - effort, phase, discipline, sprint, dates, working days, utilisation, finish
                 dates - is parsed from the plan of record, so it cannot drift from it.
    Prose      - phase names, work-item names, what each delivers, who it serves, levers,
                 milestones, dependencies, assumptions, risks - comes from the content file,
                 which is the only place it is authored.
    "Reference title" is read from the content file for the drift guard alone and is never
    written to the workbook.

TWO NUMBERS THIS WORKBOOK DELIBERATELY DOES NOT CARRY
    Effort is presented in DAYS, not hours. Hours invite a rate conversation; days are the
    planning unit. The conversion happens at write time at 8 h/day.
    The unit rate card - what a screen or an endpoint costs - is internal pricing mechanics
    and is never exported. Per-phase and per-item totals are fine; the rates are not.

FOUR GUARDS, ALL FATAL
    1. Coverage      - every work item in the plan has a content entry and vice versa, every
                       phase likewise, and no required field is empty.
    2. Drift         - the content entry's recorded reference title still matches the plan's.
                       Catches a renumbering silently moving prose onto the wrong work item.
    3. Team names    - no Nagarro-side name reaches a client-facing cell.
    4. Leakage       - every cell of the built workbook is re-read from disk and scanned for
                       file names, paths, requirement/gap/test/backlog identifiers, table
                       names, endpoints, code spans, screen numbers and machine tag paths.
                       On any hit the workbook is deleted and the build exits non-zero, so a
                       file on disk means all four passed.

REQUIREMENTS
    openpyxl (tested against 3.1.5). Nothing else.

GOTCHA
    If the output is open in Excel the write fails with PermissionError - a ~$ lock file in
    MVP-1/SRS/ is the tell. Close it and re-run.
"""
import os, re, sys

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, '..', '..', '..'))
DEV = os.path.join(ROOT, 'MVP-1', 'ProjectPlan', 'Development')

TASKS = os.path.join(DEV, 'TaskBreakdown.md')
STAFFED = os.path.join(DEV, 'StaffedSprintPlans.md')
CONTENT = os.path.join(HERE, 'DevelopmentPlanContent.md')
DEFAULT_OUT = os.path.join(ROOT, 'MVP-1', 'SRS', 'FlatWire_DevelopmentPlan.xlsx')

ISSUE_DATE = 'August 13, 2026'
HOURS_PER_DAY = 8

# Nagarro-side names. No client-facing cell may carry one.
TEAM_NAMES = ('Jaspreet', 'Srikanth', 'Shray', 'Yogender')

# Delivery order, per the plan of record's dependency chain.
PHASE_ORDER = ['1A', '1B', '1C', '3', '4', '5', '6', '7', '8', '9', '10', '11', '13', '12', '14']

# Build streams -> what a client would call them.
DISCIPLINE = {'FE': 'Operator screens', 'BE': 'Application services',
              'DB': 'Data', 'RT': 'Live data and machine interface'}


# ----------------------------------------------------------------- markdown helpers
# clean/render/_fold_blank_runs/norm are the questions-workbook helpers unchanged. The
# rich-text fold in particular is a real openpyxl 3.1.5 bug fix, not a preference.

def clean(text):
    """Strip the markdown a table cell carries."""
    text = re.sub(r'\[([^\]]*)\]\([^)]*\)', r'\1', text)   # links -> label
    text = text.replace('`', '').replace('**', '').replace('*', '')
    return re.sub(r'\s+', ' ', text).strip()


_BOLD = InlineFont(rFont='Calibri', sz=11, b=True)
_ITALIC = InlineFont(rFont='Calibri', sz=11, i=True)
_EMPHASIS = re.compile(r'\*\*(.+?)\*\*|(?<![\w*])\*([^*\n]+?)\*(?![\w*])')


def _fold_blank_runs(parts):
    """Never emit a whitespace-only run.

    openpyxl 3.1.5 writes xml:space="preserve" on a whitespace-EDGED run (' x ') but
    omits it on a whitespace-ONLY one (' '), so XML normalisation strips the content and
    Excel reports 'Repaired Records: String properties' on open. Two **bold** spans
    separated by a single space produce exactly that run.
    """
    def attach(part, ws, before=False):
        if isinstance(part, str):
            return (ws + part) if before else (part + ws)
        text = (ws + part.text) if before else (part.text + ws)
        return TextBlock(part.font, text)

    out, pending = [], ''
    for part in parts:
        if isinstance(part, str):
            part = pending + part
            pending = ''
            if part and not part.strip():
                if out:
                    out[-1] = attach(out[-1], part)
                else:
                    pending = part          # nothing to fold back into yet
                continue
            if out and isinstance(out[-1], str):
                out[-1] += part
            else:
                out.append(part)
        else:
            if pending:
                part, pending = attach(part, pending, before=True), ''
            out.append(part)
    if pending:
        out[-1] = attach(out[-1], pending) if out else pending
    return out


def render(text):
    """Turn authored markdown emphasis into Excel rich text."""
    text = text.replace('`', '')
    if not _EMPHASIS.search(text):
        return text
    parts, pos = [], 0
    for m in _EMPHASIS.finditer(text):
        if m.start() > pos:
            parts.append(text[pos:m.start()])
        if m.group(1) is not None:
            parts.append(TextBlock(_BOLD, m.group(1)))
        else:
            parts.append(TextBlock(_ITALIC, m.group(2)))
        pos = m.end()
    if pos < len(text):
        parts.append(text[pos:])
    return CellRichText(*_fold_blank_runs(parts))


def norm(text):
    """Normalise a title for the drift guard: letters, digits and single spaces only."""
    text = clean(text).lower()
    text = re.sub(r'[^a-z0-9 ]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def days(hours, dp=0):
    """Hours -> days. The workbook's only effort unit."""
    d = hours / HOURS_PER_DAY
    return f'{d:.{dp}f}' if dp else str(int(round(d)))


# Internal shorthand that is accurate but reads as jargon on a client sheet. Applied only to
# parsed commentary, never to figures - rewriting a parsed number is exactly the drift the
# parse/author split exists to prevent.
PLAIN = [(r'\bQ([1-4]) (\d{4})\b', lambda m: f'{["first","second","third","fourth"][int(m.group(1)) - 1]}'
                                             f'-quarter {m.group(2)}'),
         (r'\bQ([1-4])\b', lambda m: f'the {["first","second","third","fourth"][int(m.group(1)) - 1]} quarter'),
         (r'~', 'about ')]


def plain(text):
    for pattern, repl in PLAIN:
        text = re.sub(pattern, repl, text)
    return re.sub(r'\s+', ' ', text).strip()


# ----------------------------------------------------------------------- parsers

def parse_tasks(path):
    """Work items with development hours, from the plan's story bodies.

    QA- and BA-only stories carry no build hours. They are not development work and are
    excluded here rather than silently absorbed - which is also why this returns 107 of
    the plan's 116 stories.
    """
    text = open(path, encoding='utf-8').read()
    pat = (r'^###### (FW-[\w\d]+) · (.+?)$\n'
           r'\*\*Hours:\*\* (.+?) · \*\*Priority:\*\* (.+?) · \*\*Sprint:\*\* .+?'
           r' · \*\*Phase:\*\* (.+?) · \*\*Stream:\*\* ')
    out = {}
    for sid, title, hrs, pri, phase in re.findall(pat, text, re.M):
        streams = {}
        for m in re.finditer(r'(\d+)\s*h\s*([A-Z]{2})', hrs):
            if m.group(2) in DISCIPLINE:
                streams[m.group(2)] = streams.get(m.group(2), 0) + int(m.group(1))
        if not streams:
            continue
        out[sid] = dict(id=sid, title=' '.join(title.split()), priority=pri.strip(),
                        phase=phase.strip().replace('Phase ', ''), streams=streams)
    if not out:
        sys.exit(f'FATAL: no work items parsed from {os.path.relpath(path, ROOT)}')
    return out


def _table_after(text, heading_pattern, first_header):
    """Return the rows of the first table under a heading, as lists of cleaned cells."""
    m = re.search(heading_pattern, text, re.M)
    if not m:
        sys.exit(f'FATAL: heading not found: {heading_pattern}')
    rows, header, started = [], None, False
    for line in text[m.end():].split('\n'):
        if not line.startswith('|'):
            if started:
                break
            continue
        cells = [clean(c) for c in line.strip('|').split('|')]
        if header is None:
            if first_header not in cells:
                continue
            header = cells
            started = True
            continue
        if re.match(r'^\|[\s:|-]+\|$', line):
            continue
        rows.append(cells)
    if not rows:
        sys.exit(f'FATAL: no rows under {heading_pattern}')
    return header, rows


WORD_NUMBERS = {w: i for i, w in enumerate(
    'zero one two three four five six seven eight nine ten'.split())}

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
FIRST_YEAR = 2026        # the plan's sprints start Mon 24 Aug 2026


def _add_years(sprints):
    """Stamp a year onto each sprint's end date.

    The plan's tables carry the year on the closing row only, which is fine in a document
    read top to bottom and wrong on a milestone sheet where a single date stands alone -
    the two-developer plan finishes '08 Jan', which is not 2026. The year is recovered by
    walking the sequence and rolling over whenever the month goes backwards.
    """
    year, previous = FIRST_YEAR, -1
    for s in sprints:
        end = s['dates'].split('–')[-1].strip()
        m = re.match(r'(\d+)\s+(\w+)', end)
        if not m:
            s['end_date'] = end
            continue
        month = MONTHS.index(m.group(2)[:3]) if m.group(2)[:3] in MONTHS else previous
        if 0 <= month < previous:
            year += 1
        previous = month
        s['end_date'] = f'{m.group(1)} {m.group(2)} {year}'


def parse_staffed(path):
    """Headline, sprint schedule and story allocation, for however many team sizes the plan
    carries. Sections are located by their content rather than their number, so renumbering
    the document does not silently break the build."""
    text = open(path, encoding='utf-8').read()

    # Headline - one row per team size, and the row order defines SCENARIOS.
    _, rows = _table_after(text, r'^## \d+\. The answer.*$', 'Team')
    headline = []
    for r in rows:
        n = int(re.search(r'(\d+)', r[0]).group(1))
        headline.append(dict(devs=n, sprints=r[1], finish=r[2],
                             vs_target=r[3], vs_production=r[4]))
    scenarios = tuple(h['devs'] for h in headline)

    # Sprint schedule - one section per scenario, each headed "N developers ... finish DATE".
    # The plan spells the team size out ("Three developers"), so both forms are accepted.
    schedule, finish = {}, {}
    for m in re.finditer(r'^## \d+\. .*?(\w+) developers .*finish\s+(.+)$', text, re.M):
        word = m.group(1).lower()
        devs = int(word) if word.isdigit() else WORD_NUMBERS.get(word, 0)
        if devs not in scenarios:
            continue
        finish[devs] = clean(m.group(2))
        _, rows = _table_after(text, re.escape(m.group(0)), 'Sprint')
        parsed = [dict(sprint=r[0], dates=r[1], workdays=int(r[2]),
                       capacity=int(re.search(r'(\d+)', r[3]).group(1)),
                       planned=int(re.search(r'(\d+)', r[4]).group(1)),
                       utilisation=r[5], items=int(r[6]),
                       phases=[p.strip() for p in r[7].split(',') if p.strip()])
                  for r in rows]
        # The last row of each table is a total, not a sprint - it has no S## label. Counting
        # it as one would put the sprint count one above the headline table's own figure.
        schedule[devs] = dict(
            sprints=[s for s in parsed if re.fullmatch(r'S\d+', s['sprint'])],
            total=next((s for s in parsed if not re.fullmatch(r'S\d+', s['sprint'])), None))
        _add_years(schedule[devs]['sprints'])
    missing = [n for n in scenarios if n not in schedule]
    if missing:
        sys.exit(f'FATAL: the headline names {missing} developer(s) with no sprint table')

    # Allocation matrix - per-story effort and the sprint it lands in, one column per scenario
    # after the fixed five. The header names them, so the column order is not assumed.
    header, rows = _table_after(text, r'^## \d+\. Story.*$', 'Story')
    sprint_cols = {}
    for i, head in enumerate(header[5:], start=5):
        found = re.search(r'(\d+)', head)
        if found and int(found.group(1)) in scenarios:
            sprint_cols[int(found.group(1))] = i
        elif len(scenarios) == 1:
            sprint_cols[scenarios[0]] = i        # a single-team plan needs no label
    if sorted(sprint_cols) != sorted(scenarios):
        sys.exit(f'FATAL: allocation matrix has sprint columns for {sorted(sprint_cols)}, '
                 f'the headline names {sorted(scenarios)}')
    matrix = {}
    for r in rows:
        matrix[r[0]] = dict(id=r[0], title=r[1], phase=r[2], stream=r[3], hours=int(r[4]),
                            sprint={n: r[i] for n, i in sprint_cols.items()})
    if not matrix:
        sys.exit('FATAL: no allocation matrix rows parsed')
    return scenarios, headline, schedule, finish, matrix


PART_FIELDS = {
    'Phases': ('Reference title', 'Phase name', 'Delivers', 'Audience', 'Deferrable'),
    'Work Items': ('Reference title', 'Work item', 'Delivers', 'Audience'),
    'What changes the date': ('Lever', 'Effect', 'Detail'),
    'Milestones and dependencies': ('Item', 'Type', 'Detail', 'Covers'),
    'Assumptions and risks': ('Item', 'Type', 'Detail'),
}
OPTIONAL = {'Covers'}


def parse_content(path):
    """Parse the authored content file into {part name: {key: {field: text}}}."""
    text = open(path, encoding='utf-8').read()
    parts, current = {}, None
    known = {f for fields in PART_FIELDS.values() for f in fields}

    for chunk in re.split(r'^# Part \d+ — (.+)$', text, flags=re.M)[1:]:
        name = chunk.strip().split('\n')[0].strip()
        if name in PART_FIELDS:
            current = name
            parts[current] = {}
            continue
        if current is None:
            continue
        blocks = re.split(r'^## ([\w-]+)\s*$', chunk, flags=re.M)
        for key, body in zip(blocks[1::2], blocks[2::2]):
            entry, field = {}, None
            for line in body.split('\n'):
                if line.strip() in ('---', ''):
                    field = None if line.strip() == '---' else field
                    continue
                m = re.match(r'^\*\*([A-Z][^:*]*):\*\*\s?(.*)$', line)
                if m and m.group(1) in known:
                    field = m.group(1)
                    entry[field] = m.group(2).strip()
                elif field:
                    entry[field] += ' ' + line.strip()
            unknown = set(entry) - set(PART_FIELDS[current])
            if unknown:
                sys.exit(f'FATAL: {current} entry {key} has unknown field(s): {sorted(unknown)}')
            parts[current][key] = {k: re.sub(r'\s+', ' ', v).strip() for k, v in entry.items()}

    missing = set(PART_FIELDS) - set(parts)
    if missing:
        sys.exit(f'FATAL: content file is missing part(s): {sorted(missing)}')
    return parts


# ------------------------------------------------------------------------- guards

def guard(tasks, matrix, content):
    """Returns the deliberately-excluded story ids; exits non-zero on any real failure."""
    errors = []

    # 1/2. Coverage and drift, work items.
    #
    # The workbook's scope is the SPRINT PLAN, not the whole backlog - the sprint plan is
    # what excluded Phase 12 (yield, cost ledger, scrap) on 13 Aug 2026. So coverage is
    # checked against the allocation matrix, and a backlog story absent from it is an
    # exclusion rather than an error. It is still reported: a story that fell out by
    # accident and one that was removed by decision look identical here, and only the
    # printed list distinguishes them.
    items = content['Work Items']
    excluded = sorted(set(tasks) - set(matrix))
    for sid in sorted(set(matrix) - set(items)):
        errors.append(f'{sid} is in the sprint plan with no content entry')
    for sid in sorted(set(items) - set(matrix)):
        errors.append(f'{sid} has a content entry but no sprint allocation'
                      + (' - it is excluded from the plan' if sid in tasks else ''))
    for sid in sorted(set(matrix) - set(tasks)):
        errors.append(f'{sid} is allocated to a sprint but is not a development work item')

    for sid, entry in sorted(items.items()):
        for f in PART_FIELDS['Work Items']:
            if not entry.get(f):
                errors.append(f'{sid} is missing required field "{f}"')
        if sid not in tasks:
            continue
        a, b = norm(entry.get('Reference title', '')), norm(tasks[sid]['title'])
        n = min(20, len(a), len(b))
        if n == 0 or a[:n] != b[:n]:
            errors.append(
                f'{sid} drift: content says "{clean(entry.get("Reference title", ""))}" '
                f'but the plan says "{clean(tasks[sid]["title"])}" '
                f'- a renumbering would put prose on the wrong work item')

    # Coverage, phases - again against the sprint plan rather than the backlog.
    plan_phases = {m['phase'] for m in matrix.values()}
    phases = content['Phases']
    for p in sorted(plan_phases - set(phases)):
        errors.append(f'phase {p} carries work in the sprint plan but has no content entry')
    for p in sorted(set(phases) - plan_phases):
        errors.append(f'phase {p} has a content entry but carries no work in the sprint plan')
    for p, entry in sorted(phases.items()):
        for f in PART_FIELDS['Phases']:
            if not entry.get(f):
                errors.append(f'phase {p} is missing required field "{f}"')
    unordered = sorted(set(phases) - set(PHASE_ORDER))
    if unordered:
        errors.append(f'phase(s) {unordered} are not in the delivery order')

    # The remaining parts: required fields only.
    for part in ('What changes the date', 'Milestones and dependencies', 'Assumptions and risks'):
        if not content[part]:
            errors.append(f'part "{part}" has no entries')
        for key, entry in sorted(content[part].items()):
            for f in PART_FIELDS[part]:
                if f not in OPTIONAL and not entry.get(f):
                    errors.append(f'{part} entry {key} is missing required field "{f}"')

    # 3. Team names, everywhere in the authored prose.
    for part, entries in content.items():
        for key, entry in sorted(entries.items()):
            for f, v in entry.items():
                if f == 'Reference title':
                    continue                # never written to the workbook
                for name in TEAM_NAMES:
                    if name in v:
                        errors.append(f'{part} entry {key} field "{f}" names our own team: {name}')

    if errors:
        print('FATAL - build refused:')
        for e in errors:
            print('  *', e)
        sys.exit(1)
    return excluded


LEAKS = [
    (r'\.(?:md|html|sql|docx|xlsx|py|js|scss|css)\b', 'file name'),
    # The retired folder names are KEPT deliberately, so a stale citation still trips the
    # guard. Of the eight subject folders only Screens/ and Phases/ are listed: Business/,
    # Testing/, Operations/, Backend/, Frontend/, Database/, Development/ and Architecture/
    # are ordinary English words and would false-positive on client prose. They are covered
    # instead by the MVP-1/ and ProjectPlan/ prefixes above.
    (r'\b(?:MVP-1|MVP-2|Analysis|BaseDocuments|LatestDocument|DevelopmentPlan|'
     r'RequirementDocuments|Mockups|DBChanges|ProjectPlan|ShopfloorPlan|Screens|Phases)/',
     'repository path'),
    (r'\bFR-\d', 'requirement identifier'),
    (r'\bOI-\d', 'open-item identifier'),
    (r'\bG\d{1,2}\b', 'gap identifier'),
    (r'\bTC-\d', 'test-case identifier'),
    (r'\bFW-\d', 'backlog identifier'),
    (r'\bPLC-Q', 'PLC question identifier'),
    (r'\bPSG-', 'generation-spec identifier'),
    (r'\bD-\d{2}\b', 'decision identifier'),
    (r'\b(?:PCI|CHK|WLD|TRV|PSM|FRT|DM|SQ)\d{2,3}\b', 'requirement identifier'),
    (r'\b(?:CK|IX|UX|FK)_\w', 'constraint or index name'),
    (r'\b(?:POST|GET|PUT|PATCH|DELETE)\s+/', 'endpoint'),
    (r'`', 'code span'),
    (r'\bFL[123]\.\w', 'machine tag path'),
    (r'\bDashboard\s*\d', 'screen number'),
    (r'\b(?:RodStaging|AlloyProperty|CoilTraceability|CoilOutput|FlatWireRun|PassSchedule'
     r'|SpoolCheckin|RodCheckin|RodCheckout|WipRejection|SpcCheckpoint|SpcMeasurement'
     r'|RunReading|WeldEvent|RollOverride|DieChangeEvent|RunPauseEvent|PayoffPosition'
     r'|FlatwireQueue|planning_routings|wip_coil_orders|FlatWireDB|united_db|CommonDB)\b',
     'table or column name'),
    # Component and service names appear in most work-item titles and are the single most
    # likely thing to survive a careless edit. Not added: bare "Angular", ".NET", "SQL" -
    # those name the technology, which the client already knows and may legitimately ask
    # about; and "hub", "service", "grid" alone, which are ordinary English.
    (r'\b(?:MediatR|Dapper|SignalR|OPC UA|openpyxl|Serilog|AutoMapper|Polly)\b',
     'library or framework name'),
    (r'\b\w+(?:Service|Controller|Repository|Component|Hub|Dto|Handler)\b(?<!\bthe Service)',
     'code identifier'),
]


def scan_for_leaks(path):
    wb = load_workbook(path)
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                for pattern, label in LEAKS:
                    m = re.search(pattern, cell.value)
                    if m:
                        hits.append((ws.title, cell.coordinate, label, m.group(0),
                                     cell.value[max(0, m.start() - 40):m.end() + 40]))
    return hits


# --------------------------------------------------------------------- formatting

HEAD_FILL = PatternFill('solid', fgColor='1F3864')
TITLE_FONT = Font(name='Calibri', size=15, bold=True, color='1F3864')
SUB_FONT = Font(name='Calibri', size=10, italic=True, color='404040')
SECTION_FONT = Font(name='Calibri', size=12, bold=True, color='1F3864')
HEAD_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
BODY_FONT = Font(name='Calibri', size=11)
ALT_FILL = PatternFill('solid', fgColor='F5F7FA')
FOCUS_FILL = PatternFill('solid', fgColor='FFF7E0')
TYPE_FILL = {
    'Risk — high': PatternFill('solid', fgColor='F8CBCB'),
    'Risk — medium': PatternFill('solid', fgColor='FBE3C0'),
    'Needed from you': PatternFill('solid', fgColor='FBE3C0'),
    'External dependency': PatternFill('solid', fgColor='E4E7EA'),
    'Delivery milestone': PatternFill('solid', fgColor='D6E9D5'),
    'Assumption': PatternFill('solid', fgColor='D9E7F5'),
}
THIN = Side(style='thin', color='C8CDD3')
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
HEAD_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
CENTRE = Alignment(horizontal='center', vertical='top')


def sheet_header(ws, title, subtitle, tab=None):
    ws.sheet_view.showGridLines = False
    if tab:
        ws.sheet_properties.tabColor = tab
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 22
    ws['A2'] = subtitle
    ws['A2'].font = SUB_FONT
    ws.row_dimensions[2].height = 14


def widths(ws, columns):
    for i, (_, w) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def table(ws, start_row, columns, rows, centre_cols=(), zebra=True, focus_cols=()):
    """Write a header row plus data at start_row. Returns the row after the table."""
    for i, (head, _) in enumerate(columns, start=1):
        c = ws.cell(row=start_row, column=i, value=head)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = HEAD_ALIGN
        c.border = CELL_BORDER
    ws.row_dimensions[start_row].height = 30

    for r, values in enumerate(rows, start=start_row + 1):
        for i, value in enumerate(values, start=1):
            c = ws.cell(row=r, column=i, value=render(value) if value else None)
            c.font = BODY_FONT
            c.alignment = CENTRE if (i - 1) in centre_cols else TOP_WRAP
            c.border = CELL_BORDER
            if (i - 1) in focus_cols:
                c.fill = FOCUS_FILL
            elif zebra and (r - start_row) % 2 == 0:
                c.fill = ALT_FILL
    return start_row + len(rows) + 1


def section(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = SECTION_FONT
    return row + 1


def note(ws, row, lines):
    """Small-print notes under a table. Returns the row after them."""
    for line in lines:
        c = ws.cell(row=row, column=1, value=render(line))
        c.font = SUB_FONT
        c.alignment = Alignment(vertical='top', wrap_text=True)
        row += 1
    return row


# ------------------------------------------------------------------------- build

def build(out_path):
    tasks = parse_tasks(TASKS)
    scenarios, headline, schedule, finish, matrix = parse_staffed(STAFFED)
    content = parse_content(CONTENT)
    excluded = guard(tasks, matrix, content)

    # With one team size the column needs no label; with several it must carry one.
    def sprint_head(n):
        return 'Sprint' if len(scenarios) == 1 else f'Sprint — {n} devs'

    def reached_head(n):
        return 'Reached' if len(scenarios) == 1 else f'Reached with {n} developers'

    phases = content['Phases']
    items = content['Work Items']
    ordered_ids = sorted(items, key=lambda s: (PHASE_ORDER.index(matrix[s]['phase']),
                                               matrix[s]['id']))
    total_h = sum(matrix[s]['hours'] for s in ordered_ids)
    phase_h = {}
    phase_n = {}
    for s in ordered_ids:
        p = matrix[s]['phase']
        phase_h[p] = phase_h.get(p, 0) + matrix[s]['hours']
        phase_n[p] = phase_n.get(p, 0) + 1

    def phase_sprints(codes, devs):
        """First and last sprint the given phases occupy at a team size."""
        got = [matrix[s]['sprint'][devs] for s in ordered_ids if matrix[s]['phase'] in codes]
        if not got:
            return ''
        key = lambda x: int(x.lstrip('S'))
        lo, hi = min(got, key=key), max(got, key=key)
        return lo if lo == hi else f'{lo}–{hi}'

    def sprint_end(codes, devs):
        """Sprint label and end date of the last sprint the given phases occupy."""
        last = phase_sprints(codes, devs).split('–')[-1]
        for s in schedule[devs]['sprints']:
            if clean(s['sprint']) == last:
                return f'{last} · {s["end_date"]}'
        return ''

    wb = Workbook()

    # ---------------------------------------------------------------- Read Me
    ws = wb.active
    ws.title = 'Read Me'
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '1F3864'
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 110

    readme = [
        ('TITLE', 'Flat Wire Mill — Development Plan'),
        ('SUB', f'Issued {ISSUE_DATE} · {len(phases)} stages · {len(ordered_ids)} work items · '
                f'{days(total_h)} days of development effort'),
        ('GAP', ''),
        ('H', 'What this is'),
        ('P', 'The plan to build the flat wire manufacturing system: what will be built, in what '
              'order, how long each part takes, and what that means for the date it is finished. '
              f'It covers **{days(total_h)} days of development effort** across '
              f'**{len(ordered_ids)} work items**, grouped into **{len(phases)} stages**.'),
        ('P', 'It is written to be discussed rather than filed. The sheet we would start on is '
              '**Delivery Summary** — the date, and what would move it.'),
        ('GAP', ''),
        ('H', 'The sheets'),
        ('K', '**Delivery Summary** — when development completes, and what would change that date. '
              'Start here.'),
        ('K', f'**Plan by Stage** — the {len(phases)} stages of the build, what each one gives '
              f'you, and the sprint each lands in.'),
        ('K', '**Sprint Schedule** — every two-week sprint, with its dates, capacity and '
              'workload.'),
        ('K', '**Work Items** — the full detail. Every item of work, what it delivers, and the '
              'sprint it lands in. Filter it during the discussion.'),
        ('K', '**Milestones and Needs** — the dates we are working to, what we need from you, and '
              'what sits outside this team.'),
        ('K', '**Assumptions and Risks** — what the plan takes as given, and what could move it.'),
        ('GAP', ''),
        ('H', 'How to read the effort figures'),
        ('P', '**Effort is in days, and a day is one person working for one day.** Twenty days of '
              'effort is one person for four weeks, or four people for one week — it is a measure '
              'of size, not of elapsed time. The elapsed time is on the Delivery Summary and '
              'Sprint Schedule sheets, and it depends entirely on how many people are working.'),
        ('P', 'Sprints are **two weeks**, starting Monday 24 August 2026. **Working days are '
              'counted, not assumed** — public holidays are deducted from the sprints that '
              'contain them, which is why some sprints are shorter than others.'),
        ('GAP', ''),
        ('H', 'What these figures do not include'),
        ('K', '**Coil yield, costing and scrap.** This work is out of the plan by agreement and no '
              'part of it is in any figure on these sheets. Yield would continue to be reported '
              'the way it is today, and flat wire scrap routed by hand, until it is picked up.'),
        ('K', '**Testing and business analysis.** The figures are development only. Quality '
              'assurance, business analysis and contingency are planned separately and are '
              'additional to every date shown here.'),
        ('K', '**Acceptance testing and machine commissioning.** These follow the development '
              'completion date; they are not inside it.'),
        ('K', '**Two design decisions still open.** An allowance is held for each, and neither is '
              'included in any figure on these sheets. Both are on the Assumptions and Risks '
              'sheet.'),
        ('GAP', ''),
        ('H', 'What we need back'),
        ('K', 'Confirmation of the team size the plan is built on, and of the completion date it '
              'produces. The Delivery Summary sheet sets out what would move that date.'),
        ('K', 'The four items on the Milestones and Needs sheet marked as needed from you. Two of '
              'them gate work that starts early.'),
        ('K', 'Confirmation that the stage order matches your operational priorities. If something '
              'matters more than its position suggests, it can be moved.'),
    ]

    r = 1
    for kind, text in readme:
        if kind == 'GAP':
            r += 1
            continue
        cell = ws.cell(row=r, column=1 if kind in ('TITLE', 'SUB', 'H') else 2)
        if kind == 'TITLE':
            cell.value = text; cell.font = TITLE_FONT; ws.row_dimensions[r].height = 22
        elif kind == 'SUB':
            cell.value = text; cell.font = SUB_FONT
        elif kind == 'H':
            cell.value = text; cell.font = SECTION_FONT
        else:
            cell.value = render(text)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        r += 1

    # ------------------------------------------------------- Delivery Options
    ws = wb.create_sheet('Delivery Summary')
    cols = [('Team size', 16), ('Sprints', 10), ('Development complete', 22),
            ('Against the end-September target', 26), ('Against fourth-quarter production', 32)]
    plural = 'the options below are' if len(scenarios) > 1 else 'the plan below is'
    sheet_header(ws, 'Delivery Summary',
                 f'{days(total_h)} days of development effort across {len(ordered_ids)} work '
                 f'items — {plural} what that delivers, and when.',
                 tab='B71C1C')
    widths(ws, cols)
    rows = [[f'{h["devs"]} developers', h['sprints'], h['finish'],
             plain(h['vs_target']), plain(h['vs_production'])] for h in headline]
    r = table(ws, 4, cols, rows, centre_cols=(1,), focus_cols=(2,))

    r += 1
    r = section(ws, r, 'What changes the date')
    lever_cols = [('Lever', 46), ('Effect on the date', 40), ('What it means', 100)]
    lrows = [[content['What changes the date'][k]['Lever'],
              content['What changes the date'][k]['Effect'],
              content['What changes the date'][k]['Detail']]
             for k in sorted(content['What changes the date'])]
    r = table(ws, r, lever_cols, lrows)

    r = note(ws, r + 1, [
        'This is a development completion date. Acceptance testing and machine commissioning '
        'follow it and are not included in it.',
        'Coil yield, costing and scrap are out of the plan by agreement and are in none of these '
        'figures.',
    ])

    # ---------------------------------------------------------- Plan by Stage
    ws = wb.create_sheet('Plan by Stage')
    cols = [('#', 5), ('Stage', 34), ('What it delivers', 78), ('Who uses it', 26),
            ('Effort (days)', 12), ('Work items', 11)]
    cols += [(sprint_head(n), 14) for n in scenarios]
    cols += [('Can it be deferred?', 44)]
    sheet_header(ws, 'Plan by Stage',
                 f'The {len(phases)} stages of the build, in delivery order. Each row is a block '
                 f'of capability that lands together.', tab='2E7D32')
    widths(ws, cols)
    rows = []
    for i, p in enumerate([p for p in PHASE_ORDER if p in phases], start=1):
        e = phases[p]
        rows.append([str(i), e['Phase name'], e['Delivers'], e['Audience'],
                     days(phase_h.get(p, 0)), str(phase_n.get(p, 0))]
                    + [phase_sprints({p}, n) for n in scenarios]
                    + [e['Deferrable']])
    last = table(ws, 4, cols, rows, centre_cols=(0, 4, 5, 6, 7, 8))
    ws.auto_filter.ref = f'A4:{get_column_letter(len(cols))}{last - 1}'
    ws.freeze_panes = 'C5'
    note(ws, last + 1,
         [f'Stage effort is rounded to whole days, so the column sums to within a day of the '
          f'{days(total_h)}-day total rather than exactly to it.',
          'A sprint range means the stage spans those sprints; it does not occupy them '
          'exclusively — several stages run alongside each other.'])

    # -------------------------------------------------------- Sprint Schedule
    ws = wb.create_sheet('Sprint Schedule')
    sheet_header(ws, 'Sprint Schedule',
                 'Two-week sprints from Monday 24 August 2026. Working days are counted, not '
                 'assumed — public holidays are deducted from the sprints containing them, which '
                 'is why some sprints are shorter.', tab='1F3864')
    cols = [('Sprint', 10), ('Dates', 20), ('Working days', 12), ('Team capacity (days)', 16),
            ('Planned work (days)', 16), ('Utilisation', 12), ('Work items', 11),
            ('What is delivered', 86)]
    widths(ws, cols)
    r = 4
    for n in scenarios:
        r = section(ws, r, f'{n} developers — {len(schedule[n]["sprints"])} sprints, '
                           f'development complete {finish[n]}')
        rows = []
        for s in schedule[n]['sprints'] + [schedule[n]['total']]:
            if s is None:
                continue
            names = ' · '.join(phases[p]['Phase name'] for p in s['phases'] if p in phases)
            label = clean(s['sprint']) or 'Total'
            rows.append([label, s['dates'], str(s['workdays']),
                         days(s['capacity']), days(s['planned']), s['utilisation'],
                         str(s['items']), names or 'the whole plan'])
        r = table(ws, r, cols, rows, centre_cols=(0, 2, 3, 4, 5, 6)) + 1
    r = note(ws, r, [
        'Team capacity is the whole team for the whole sprint; planned work is the effort of '
        'the items placed in it. Utilisation is one against the other.',
        'Above 100 % means the sprint is planned slightly over capacity and absorbs the '
        'difference from the sprint before or after it.',
        'A short final sprint is the tail of the plan, not spare capacity — the work left in it '
        'cannot start any earlier.',
    ])

    # ------------------------------------------------------------- Work Items
    ws = wb.create_sheet('Work Items')
    cols = [('#', 6), ('Work item', 46), ('What it delivers for you', 76), ('Stage', 30),
            ('Who it serves', 26), ('Area of the system', 30), ('Effort (days)', 12)]
    cols += [(sprint_head(n), 14) for n in scenarios]
    sheet_header(ws, 'Work Items',
                 f'All {len(ordered_ids)} items of development work, in the order they are built. '
                 f'Use the filters on the header row to narrow to a stage or a sprint.',
                 tab='7B1FA2')
    widths(ws, cols)
    rows = []
    for i, sid in enumerate(ordered_ids, start=1):
        e, m = items[sid], matrix[sid]
        disciplines = ' + '.join(DISCIPLINE[s] for s in sorted(tasks[sid]['streams'],
                                                               key=lambda x: -tasks[sid]['streams'][x]))
        rows.append([str(i), e['Work item'], e['Delivers'],
                     phases[m['phase']]['Phase name'], e['Audience'], disciplines,
                     days(m['hours'], dp=1)]
                    + [m['sprint'][n] for n in scenarios])
    last = table(ws, 4, cols, rows, centre_cols=(0, 6, 7, 8, 9))
    ws.auto_filter.ref = f'A4:{get_column_letter(len(cols))}{last - 1}'
    ws.freeze_panes = 'C5'

    # -------------------------------------------------- Milestones and Needs
    ws = wb.create_sheet('Milestones and Needs')
    sheet_header(ws, 'Milestones and Needs',
                 'The dates the plan works to, what we need from you to hit them, and what sits '
                 'outside this team.', tab='EF6C00')
    md = content['Milestones and dependencies']
    cols = [('Milestone', 34)]
    cols += [(reached_head(n), 21) for n in scenarios]
    cols += [('What it means', 100)]
    widths(ws, cols)
    r = section(ws, 4, 'Delivery milestones — the end of the sprint that completes them')
    rows = []
    for k in sorted(k for k in md if md[k]['Type'] == 'Delivery milestone'):
        e = md[k]
        codes = {c.strip() for c in e.get('Covers', '').split(',') if c.strip()}
        rows.append([e['Item']]
                    + [sprint_end(codes, n) if codes else '' for n in scenarios]
                    + [e['Detail']])
    r = table(ws, r, cols, rows, centre_cols=(1, 2, 3)) + 1
    r = note(ws, r, ['Stages overlap, so two milestones can land in the same sprint — the plan '
                     'does not idle between them.']) + 1

    r = section(ws, r, 'What we need from you, and what sits outside this team')
    dcols = [('Item', 34), ('Type', 22), ('What it means', 128)]
    drows = []
    for k in sorted(k for k in md if md[k]['Type'] != 'Delivery milestone'):
        e = md[k]
        drows.append([e['Item'], e['Type'], e['Detail']])
    r = table(ws, r, dcols, drows)
    for i, values in enumerate(drows):
        cell = ws.cell(row=r - len(drows) - 1 + i + 1, column=2)
        if values[1] in TYPE_FILL:
            cell.fill = TYPE_FILL[values[1]]

    # -------------------------------------------------- Assumptions and Risks
    ws = wb.create_sheet('Assumptions and Risks')
    sheet_header(ws, 'Assumptions and Risks',
                 'What the plan takes as given, and what could move it. Nothing here is hidden in '
                 'a figure on another sheet.', tab='546E7A')
    cols = [('Item', 38), ('Type', 20), ('Detail', 128)]
    widths(ws, cols)
    ar = content['Assumptions and risks']
    order = {'Assumption': 0, 'Risk — high': 1, 'Risk — medium': 2}
    keys = sorted(ar, key=lambda k: (order.get(ar[k]['Type'], 9), k))
    rows = [[ar[k]['Item'], ar[k]['Type'], ar[k]['Detail']] for k in keys]
    last = table(ws, 4, cols, rows)
    for i, values in enumerate(rows):
        if values[1] in TYPE_FILL:
            ws.cell(row=5 + i, column=2).fill = TYPE_FILL[values[1]]
    ws.auto_filter.ref = f'A4:C{last - 1}'
    ws.freeze_panes = 'A5'

    wb.save(out_path)
    return dict(phases=len(phases), items=len(ordered_ids), days=days(total_h),
                scenarios=scenarios, excluded=excluded,
                sheets=[w.title for w in wb.worksheets])


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT
    stats = build(out)
    rel = os.path.relpath(out, ROOT)
    print(f'wrote {rel}')
    print(f'  team sizes           : {" · ".join(f"{n} developers" for n in stats["scenarios"])}')
    print(f'  stages               : {stats["phases"]}')
    print(f'  work items           : {stats["items"]}')
    print(f'  development effort   : {stats["days"]} days')
    if stats['excluded']:
        print(f'  excluded from plan   : {len(stats["excluded"])} backlog stories '
              f'({", ".join(stats["excluded"])}) - not allocated to any sprint')
    print(f'  sheets               : {" · ".join(stats["sheets"])}')

    hits = scan_for_leaks(out)
    if hits:
        print(f'\nFATAL - {len(hits)} leak(s) found in the built workbook:')
        for sheet, coord, label, found, context in hits[:40]:
            print(f'  {sheet}!{coord}  {label}: {found!r}')
            print(f'      ...{context}...')
        os.remove(out)
        sys.exit(1)
    print('  leakage scan         : clean')


if __name__ == '__main__':
    main()
