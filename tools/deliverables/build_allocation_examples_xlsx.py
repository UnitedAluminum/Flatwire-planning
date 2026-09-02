"""Render AllocationExamplesContent.md to MVP-1/SRS/FlatWire_OrderAllocationExamples.xlsx.

    python MVP-1/ProjectPlan/Tools/build_allocation_examples_xlsx.py [out.xlsx]

The client-facing workbook of rod-to-order worked examples. Its subject matter is
LatestDocument/RodOrderAllocation_WorkedExamples.md - the internal trace, which names tables,
columns and register ids. This generator renders the CLIENT wording, which names none of them.

    Structure and prose  - AllocationExamplesContent.md, the only place either is authored.
    Subject matter       - RodOrderAllocation_WorkedExamples.md (internal; not read by this script).

HISTORY, because it explains why this file looks recovered rather than designed
    This generator and its content file were documented in Tools/README.md from 24 Aug 2026 but
    were never committed - only __pycache__/build_allocation_examples_xlsx.cpython-314.pyc and the
    shipped .xlsx survived, and __pycache__ is gitignored, so the source was unrecoverable from
    git. On 25 Aug 2026 the content was extracted back out of the shipped workbook and this
    renderer written against it, so the output matches what the client already holds. That is why
    the content file says "recovered, not re-authored".

GUARDS, ALL FATAL
    1. Coverage  - every sheet named in the content file is written, and no sheet is empty.
    2. Arithmetic - every row whose cells are all numeric-or-blank is checked for a total in the
                    last column, where the content file marks one with '=sum'.
    3. Drift     - the sheet set and order match SHEET_ORDER below.
    4. Team names - no Nagarro name reaches a client cell.
    5. Leakage   - no file name, path, requirement/gap/test/story id, table or column name,
                   endpoint, constraint name or machine tag path reaches a client cell.

Needs only openpyxl.
"""
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
def _repo_root():
    """Walk up to the directory holding .git.

    Deliberately NOT a level count. These scripts have moved twice and a hard-coded
    depth broke them both times; a marker survives the next move too.
    """
    d = os.path.abspath(os.path.dirname(__file__))
    while d != os.path.dirname(d):
        if os.path.isdir(os.path.join(d, '.git')):
            return d
        d = os.path.dirname(d)
    raise SystemExit('cannot locate repository root (no .git found above %s)' % __file__)


REPO = _repo_root()
CONTENT = os.path.join(HERE, 'AllocationExamplesContent.md')
DEFAULT_OUT = os.path.join(REPO, 'deliverables', 'FlatWire_OrderAllocationExamples.xlsx')

NAVY = '1F3864'
GREY = '404040'

# Guard 3: the sheet set and their order are part of the deliverable's contract - the client reads
# them in this sequence and the Read Me sheet tells them to.
SHEET_ORDER = [
    'Read Me', 'The Rules', 'Worked Examples', 'One Rod, One Order', 'Two Rods, One Spool',
    'Order Handoff', 'Processing Order Options', 'Weight and Footage', 'Order Completion',
    'Full Planned Run', 'Heavier Rod Run', 'Orders Across the Run', 'What We Need Confirmed',
]

# Guard 4. Ours, not theirs - a client cell must never name our team.
OUR_TEAM = ['Srikanth', 'Yogender', 'Ritika', 'Ashwani', 'Shray', 'Waseem', 'Vicky', 'Divesh',
            'Sushant', 'Nagarro']

# Guard 5. Kept in step with build_questions_xlsx.py's LEAKS list; see the note there about why
# the bare words Spool, Stand, Drawer, Die, Edger and Dancer are deliberately absent - they are
# what operators call the physical articles and would false-positive on ordinary client prose.
# The die split (2 Sep 2026) added ToolingInventoryDie and DieHistory; both are compounds and so
# safe to guard. DIE IS WORSE THAN DRAWER - "die"/"dies" is constant in client prose about
# tooling - so it must never be added.
LEAKS = [
    (r'\.(?:md|html|sql|docx|xlsx|py|js|scss|css)\b', 'file name'),
    (r'\b(?:00-overview|10-requirements|20-architecture|30-database|40-backend|50-frontend|60-delivery|70-testing|80-operations|90-registers|95-archive|'
     r'deliverables|tools|screens|phases|tasks|mockups|schema|sql|scripts)/',
     'repository path'),
    (r'\bFR-\d', 'requirement identifier'),
    (r'\bOI-\d', 'open-item identifier'),
    (r'\bG\d{1,2}\b', 'gap identifier'),
    (r'\bTC-\d', 'test-case identifier'),
    (r'\bFW-\d{3}\b', 'backlog identifier'),
    (r'\bORD\d{3}\b', 'requirement identifier'),
    # NOT guarded: bare Q## question numbers. They are the client's own cross-reference into
    # FlatWire_ClientQuestions.xlsx - that workbook exists so the client can answer them by
    # number - so they are shared vocabulary, not leakage. build_questions_xlsx.py omits them for
    # the same reason. PLC-Q and PSG- ARE internal series and stay guarded.
    (r'\bPLC-Q', 'PLC question identifier'),
    (r'\bD-\d{2}\b', 'decision identifier'),
    (r'\b(?:CK|IX|UX|FK)_\w', 'constraint or index name'),
    (r'\b(?:POST|GET|PUT|PATCH|DELETE)\s+/', 'endpoint'),
    (r'`', 'code span'),
    (r'\bFL[123]\.\w', 'machine tag path'),
    (r'\b(?:RodOrderAllocation|RodOrderConsumption|SpoolProcessing|SpoolStaging|SpoolOrder'
     r'|SpoolTraceability|SpoolConfiguration|SpoolCarrier|CoilTraceability|CoilOutput'
     r'|FlatWireRun|FlatWireRunDetail|RodStaging|RodCheckin|RodCheckout|SpoolCheckin'
     r'|PassSchedule|AlloyProperty|PayoffPosition|FlatWireDB|united_db|CommonDB'
     r'|ToolingInventoryDie|DieHistory'
     r'|planning_routings|wip_coil_orders)\b', 'table or column name'),
    (r'\bPinRole|PinnedFirst|PinnedLast|PinnedBoth\b', 'column value'),
]

KINDS = ('title', 'subtitle', 'section', 'header', 'data', 'blank')


def die(msg, extra=()):
    print('FATAL - ' + msg)
    for line in extra:
        print('  * ' + line)
    sys.exit(1)


# --------------------------------------------------------------------------- parse

def parse(path):
    """-> [ {name, widths: {col: width}, freeze: str|None, rows: [(kind, [cells])]} ]"""
    if not os.path.exists(path):
        die('content file not found: %s' % path)
    sheets, cur = [], None
    for raw in open(path, encoding='utf-8').read().split('\n'):
        line = raw.rstrip()
        m = re.match(r'^## Sheet:\s*(.+?)\s*$', line)
        if m:
            cur = {'name': m.group(1), 'widths': {}, 'freeze': None, 'filter': None, 'rows': []}
            sheets.append(cur)
            continue
        if cur is None:
            continue
        if line.startswith('!widths '):
            for pair in line[len('!widths '):].split():
                col, _, w = pair.partition('=')
                cur['widths'][col] = float(w)
            continue
        if line.startswith('!freeze '):
            cur['freeze'] = line[len('!freeze '):].strip()
            continue
        if line.startswith('!filter '):
            cur['filter'] = line[len('!filter '):].strip()
            continue
        m = re.match(r'^(%s):\s?(.*)$' % '|'.join(KINDS), line)
        if not m:
            continue
        kind, body = m.group(1), m.group(2)
        if kind == 'blank':
            cur['rows'].append(('blank', []))
            continue
        cells = [c.strip().replace('\\|', '|').replace('<br>', '\n')
                 for c in body.split(' | ')] if body else ['']
        cur['rows'].append((kind, cells))
    if not sheets:
        die('content file defines no sheets')
    return sheets


def as_number(text):
    """Return a float for a plain numeric cell, else None. Commas and units are not numbers.

    A leading apostrophe forces text, exactly as it does in Excel. The workbook needs this: it
    writes counts and spool numbers as real numbers, but weights and pounds-per-foot factors as
    TEXT so their significant figures survive -- 0.080850 must not render as 0.08085.
    """
    if text.startswith("'"):
        return None
    t = text.strip().replace(',', '')
    if not t or not re.fullmatch(r'-?\d+(?:\.\d+)?', t):
        return None
    return float(t)


def as_text(text):
    """The literal string to write when the cell is not a number: drop the forcing apostrophe."""
    return text[1:] if text.startswith("'") else text


# --------------------------------------------------------------------------- guards

def guard_structure(sheets):
    names = [s['name'] for s in sheets]
    errors = []
    if names != SHEET_ORDER:
        missing = [n for n in SHEET_ORDER if n not in names]
        extra = [n for n in names if n not in SHEET_ORDER]
        for n in missing:
            errors.append('sheet missing from the content file: "%s"' % n)
        for n in extra:
            errors.append('sheet not in SHEET_ORDER: "%s"' % n)
        if not missing and not extra:
            errors.append('sheet order differs from SHEET_ORDER:\n      content: %s\n      expected: %s'
                          % (names, SHEET_ORDER))
    for s in sheets:
        if not any(k != 'blank' for k, _ in s['rows']):
            errors.append('sheet "%s" has no content rows' % s['name'])
    if errors:
        die('%d structure problem(s):' % len(errors), errors)


def guard_arithmetic(sheets):
    """Where a data row's last cell is marked '=sum', the preceding numbers must add up to it.

    This is the guard the other generators do not have: these figures are COMPUTED, so a typo in
    the content file is a wrong number in front of the client rather than a missing sentence.
    """
    errors = []
    for s in sheets:
        for i, (kind, cells) in enumerate(s['rows'], start=1):
            if kind != 'data' or len(cells) < 3:
                continue
            if not cells[-1].endswith('=sum'):
                continue
            stated = as_number(cells[-1][:-4])
            parts = [as_number(c) for c in cells[1:-1]]
            if stated is None or any(p is None for p in parts):
                errors.append('%s row %d: "=sum" needs a numeric total and numeric parts'
                              % (s['name'], i))
                continue
            total = sum(parts)
            if abs(total - stated) > 0.005:
                errors.append('%s row %d: parts sum to %g but the row states %g'
                              % (s['name'], i, total, stated))
    if errors:
        die('%d arithmetic problem(s):' % len(errors), errors)


def guard_text(sheets):
    errors = []
    for s in sheets:
        for i, (kind, cells) in enumerate(s['rows'], start=1):
            for cell in [as_text(x) for x in cells]:
                for name in OUR_TEAM:
                    if re.search(r'\b%s\b' % re.escape(name), cell):
                        errors.append('%s row %d names our own team: %s' % (s['name'], i, name))
                for pattern, label in LEAKS:
                    m = re.search(pattern, cell)
                    if m:
                        errors.append('%s row %d leaks a %s: %r'
                                      % (s['name'], i, label, m.group(0)))
    if errors:
        die('%d leak(s) / team name(s) in client cells:' % len(errors), errors)


# --------------------------------------------------------------------------- render

def render(sheets, out):
    wb = Workbook()
    wb.remove(wb.active)
    title_font = Font(bold=True, size=15, color=NAVY)
    sub_font = Font(size=10, color=GREY)
    sect_font = Font(bold=True, size=12, color=NAVY)
    head_font = Font(bold=True, size=11, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor=NAVY)
    head_align = Alignment(wrap_text=True, vertical='center')
    data_align = Alignment(wrap_text=True, vertical='top')

    for s in sheets:
        ws = wb.create_sheet(s['name'])
        r = 0
        for kind, cells in s['rows']:
            r += 1
            if kind == 'blank':
                continue
            for c, text in enumerate(cells, start=1):
                if text.endswith('=sum'):
                    text = text[:-4]
                num = as_number(text)
                cell = ws.cell(row=r, column=c,
                               value=num if num is not None else as_text(text))
                if kind == 'title':
                    cell.font = title_font
                elif kind == 'subtitle':
                    cell.font = sub_font
                elif kind == 'section':
                    cell.font = sect_font
                elif kind == 'header':
                    cell.font, cell.fill, cell.alignment = head_font, head_fill, head_align
                else:
                    cell.alignment = data_align
            if kind == 'title':
                ws.row_dimensions[r].height = 22
            elif kind == 'header':
                ws.row_dimensions[r].height = 30
        for col, width in s['widths'].items():
            ws.column_dimensions[col].width = width
        if s['freeze']:
            ws.freeze_panes = s['freeze']
        # The autofilter range is CARRIED from the content file, not derived. Deriving it from
        # the header row and max_row got two sheets wrong: 'Read Me' has no filter at all, and
        # 'Full Planned Run' stops at its total row rather than the last populated row.
        if s['filter']:
            ws.auto_filter.ref = s['filter']
    wb.save(out)
    return wb


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT
    sheets = parse(CONTENT)
    guard_structure(sheets)
    guard_arithmetic(sheets)
    guard_text(sheets)
    wb = render(sheets, out)
    rel = os.path.relpath(out, REPO)
    print('wrote %s' % rel)
    print('  sheets               : %d' % len(wb.worksheets))
    print('  rows                 : %d' % sum(len(s['rows']) for s in sheets))
    print('  structure            : sheet set and order match')
    print('  arithmetic           : every marked total adds up')
    print('  team names           : clean')
    print('  leakage scan         : clean')


if __name__ == '__main__':
    main()
