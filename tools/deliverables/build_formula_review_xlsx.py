"""
Build the client-facing review workbook for United Aluminum's pass-calculator formulas.

    python tools/deliverables/build_formula_review_xlsx.py [output.xlsx]

With no arguments it writes deliverables/FlatWire_FormulaReview.xlsx.

WHY THIS EXISTS
    Tim O'Brien sent twenty formulas on 2 Sep 2026 as a Word document in which every
    formula is a PNG image - there is no selectable formula text anywhere in the file.
    Reading them is therefore interpretation, not extraction, and seven readings could
    not be settled from the page. This workbook is how those go back to the client.

    It does two things an e-mail cannot. Sheet 2 hands our transcription back in text,
    so the client can check what we read against what he wrote - which is the root
    issue rather than a side effect of it. Sheets 3 and 4 are LIVE: every output cell
    is an Excel formula over named inputs, so the client can change an input and watch
    our implementation respond. If we have mis-read a formula, that is where it shows.

WHAT IS DELIBERATE
    Sheet 3 will show #NUM! when final thickness reaches half the entry diameter. That
    is not a defect in the sheet - it is the domain limit raised in point 3, reproduced
    rather than described.

    Sheet 4's SUM column is a live formula, not a typed zero. Changing any width leaves
    it at zero, which is the whole of point 2.

LEAKAGE RULE
    This is a client deliverable. It carries no internal identifiers - no requirement,
    gap, question or story numbers, no table or file names, no paths. Only the client's
    own formula headings and plain language. A guard at the end enforces it and is fatal.

REQUIREMENTS
    openpyxl (tested against 3.1.5). Nothing else.

GOTCHA
    If the output is open in Excel the write fails with PermissionError - a ~$ lock file
    in deliverables/ is the tell. Close it and re-run.
"""
import math
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- palette, matching the questions workbook -------------------------------
HEAD_FILL = PatternFill('solid', fgColor='1F3864')
TITLE_FONT = Font(name='Calibri', size=15, bold=True, color='1F3864')
SUB_FONT = Font(name='Calibri', size=10, italic=True, color='404040')
HEAD_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
BODY_FONT = Font(name='Calibri', size=11)
MONO_FONT = Font(name='Consolas', size=10)
BOLD_BODY = Font(name='Calibri', size=11, bold=True)
RESP_FILL = PatternFill('solid', fgColor='FFF7E0')
ALT_FILL = PatternFill('solid', fgColor='F5F7FA')
INPUT_FILL = PatternFill('solid', fgColor='E2EFDA')
BAD_FILL = PatternFill('solid', fgColor='F8CBCB')
OK_FILL = PatternFill('solid', fgColor='E2EFDA')
PRIORITY = {'Answer needed': PatternFill('solid', fgColor='FBE3C0'),
            'Calculator will show': PatternFill('solid', fgColor='D9E7F5')}
TOP_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
HEAD_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sheet(wb, title, subtitle, columns, first_response=None):
    """columns: list of (header, width). Returns (ws, header_row)."""
    ws = wb.create_sheet(title)
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws['A2'] = subtitle
    ws['A2'].font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(2, len(columns)))
    hr = 4
    for i, (head, width) in enumerate(columns, start=1):
        c = ws.cell(row=hr, column=i, value=head)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = HEAD_ALIGN
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[hr].height = 30
    ws.freeze_panes = f'A{hr + 1}'
    return ws, hr


def put(ws, r, values, fills=None, font=None, height=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = font or BODY_FONT
        c.alignment = TOP_WRAP
        c.border = BOX
        if fills and fills.get(i):
            c.fill = fills[i]
    if height:
        ws.row_dimensions[r].height = height


# ---- the client's formulas, as transcribed ----------------------------------
PI = math.pi


def f16(D1, T2, R, C5=1.0):
    b = 1 - 2 * T2 / D1
    if b <= 0:
        return None
    return C5 * ((0.7854 * D1 ** 2 / T2) * (1 - 15.8 * b ** 2.25 * ((2 * R / D1) ** -0.82))
                 + 0.1426 * D1 * (2 * T2 / D1))


def solve_D1(T2, w_target, R, C5=1.0):
    lo, hi = T2 * 2.000001, 1.0
    for _ in range(300):
        m = (lo + hi) / 2
        v = f16(m, T2, R, C5)
        if v is None or v < w_target:
            lo = m
        else:
            hi = m
    return (lo + hi) / 2


D1 = solve_D1(0.085, 0.700, 6.000)          # 0.2842 - entry to the 12" mill
AE1 = PI * D1 ** 2 / 4
ROUNDED = 0.700 * 0.085 - (0.085 ** 2 - PI * 0.085 ** 2 / 4)
RECT = 0.700 * 0.085

wb = Workbook()
wb.remove(wb.active)

# =============================================================== SHEET 1 =====
POINTS = [
    (1, 'Linear Feet Round to Flat  /  Linear Feet (Flat to Flat)  vs  the two area formulas',
     'Answer needed',
     'The two linear-feet formulas use width x thickness. The two area formulas use the '
     'rectangle-with-rounded-edges. The document does not say which applies when footage is '
     'calculated.',
     f'At the 12" mill, 0.085t x 0.700w:\n'
     f'  entry area          = {AE1:.6f} in²\n'
     f'  exit, rounded edge  = {ROUNDED:.6f} in²\n'
     f'  exit, width x thick = {RECT:.6f} in²\n'
     f'  elongation, rounded = {AE1/ROUNDED:.3f}  (+9.4%)\n'
     f'  elongation, w x t   = {AE1/RECT:.3f}  (+6.6%)\n'
     'On 1,000 ft in: 1,094 ft against 1,066 ft.',
     'A 2.7% difference in delivered footage on a single pass. The correction is 0.2146 x '
     'thickness / width, so it is 2.6% at your 0.085 x 0.700 output and 0.5% by the 0.016 x '
     '0.699 finish. On a thicker, narrower section - 0.110 x 0.625 - it reaches 3.8%, and there '
     'the width x thickness basis returns an elongation of 0.98: the wire leaving the mill '
     'shorter than it entered, which a reduction pass cannot do.',
     'Should footage be computed on the rounded-edge area rather than width x thickness?'),

    (2, 'Cumulative True Strain', 'Answer needed',
     'The formula adds the length, width and thickness strains. Summed across the three '
     'dimensions of one pass, that quantity is always zero for any pass that conserves volume.',
     'Substituting your own "Linear Feet (Flat to Flat)" makes it exact: it gives\n'
     '  L2 x W2 x T2 = L1 x W1 x T1\n'
     'so the sum is ln(1) = 0 for every flat pass, whatever the numbers. See the "Strain check" '
     'sheet, where the sum is a live formula.\n\n'
     'The round-to-flat pass does not return zero - it returns ln(pi/4) = -0.2416, the same '
     'value for every product, because all that survives is the 4/pi between a round section\'s '
     'width x thickness and its true area.',
     'Accumulated cold work is what decides whether a product can run the hybrid route from the '
     '12" mill straight into the finishing mill, which carries no intermediate anneal. A figure '
     'reading zero - or negative, as the whole-line total does - would clear material that '
     'cannot tolerate it.',
     'Do you intend a sum across successive passes rather than across the three dimensions of '
     'one pass?'),

    (3, 'Calculated Width Round to Flat', 'Answer needed',
     'The term (1 - 2T2/D1) raised to 2.25 has no real value once final thickness reaches half '
     'the entry diameter. Above that there is a further band where the predicted width implies '
     'the section grew.',
     'For a 0.2842 entry at the 12" mill, undefined at T2 >= 0.1421 (any reduction under 50%). '
     'Then:\n'
     '  T2 0.140  50.7%  width 0.4929  elongation 0.979  impossible\n'
     '  T2 0.130  54.3%  width 0.5235  elongation 0.984  impossible\n'
     '  T2 0.120  57.8%  width 0.5568  elongation 0.995  impossible\n'
     '  T2 0.115  59.5%  width 0.5746  elongation 1.003  ok\n'
     '  T2 0.110  61.3%  width 0.5931  elongation 1.012  ok\n'
     '  T2 0.085  70.1%  width 0.7000  elongation 1.094  your output',
     'The usable floor sits near 59% reduction on this entry. Your 0.085 output is comfortably '
     'clear of it. We would rather set the limits to your practice than guess where they belong.',
     'What is the least reduction you would take on the flattening pass in practice?'),

    (4, 'Final Cross-Sectional Area Flat Wire', 'Calculator will show',
     'As printed the formula subtracts an area from a width, which cannot be evaluated as it '
     'stands. We read a missing multiplication by final thickness.',
     'At 0.085 x 0.700:\n'
     f'  as printed        0.700 - 0.001550         = {0.700 - 0.001550:.6f}\n'
     f'  with x thickness  0.700 x 0.085 - 0.001550 = {ROUNDED:.6f} in²\n'
     'A factor of 12 apart, and roughly 1 / thickness - so it worsens as gauge falls. With the '
     'multiplication restored it matches your "Round to Flat Cross-Sectional Area" exactly.',
     'Only the corrected form is usable, and it agrees with your other area formula to twelve '
     'decimal places - which is why we are confident it is a transcription slip rather than a '
     'difference of method.',
     'Is the missing multiplication by final thickness the intended reading?'),

    (5, 'Round to Flat Cross-Sectional Area  /  Calculated Width Round to Flat',
     'Calculator will show',
     'Two symbols appear to serve two purposes each. The legend gives r as the wire rod radius, '
     'and D1 as the original diameter, but neither reading works in every formula that uses them.',
     'r, on the 0.085 x 0.700 section:\n'
     f'  r = rod radius 0.1421   correction 0.017329  ->  area {0.0595 - (D1/2)**2*(4-PI):.6f} in²\n'
     f'  r = edge radius 0.0425  correction 0.001550  ->  area {ROUNDED:.6f} in²\n'
     '27.2% apart. Only the edge radius reproduces the rounded-edge area. Your legend is correct '
     'for the two linear-feet formulas, where r genuinely is the rod radius.\n\n'
     'D1 is plainly the incoming rod in "Linear Feet Round to Round" and "Area Reduction '
     'Percentage". In "Calculated Width Round to Flat" it has to be the diameter entering the '
     'flattening pass, after the draw boxes.',
     'Reading r as the rod radius understates the finished area by more than a quarter, and the '
     'error grows as the square of the flattening ratio. It is dimensionally valid either way, so '
     'nothing catches it automatically.',
     'Can you confirm r is the edge radius in the area formula, and that D1 is the flattening '
     'entry diameter in the width formula?'),

    (6, 'The three empirical factors', 'Calculator will show',
     'The spread factors and the edging factor are placeholders pending sample runs. We do not '
     'know whether they are single numbers or tables, nor what a table would be looked up by.',
     'Sensitivity of the round-to-flat factor at the 12" mill, 0.2842 entry to 0.085:\n'
     f'  factor 0.95  ->  width {f16(D1,0.085,6.0,0.95):.4f}\n'
     f'  factor 1.00  ->  width {f16(D1,0.085,6.0,1.00):.4f}\n'
     f'  factor 1.05  ->  width {f16(D1,0.085,6.0,1.05):.4f}\n'
     f'  factor 1.10  ->  width {f16(D1,0.085,6.0,1.10):.4f}\n'
     'A 5% move shifts width by 0.035 in - seven times the +/-0.005 tolerance band.\n\n'
     'The flat-to-flat factor is not on the same scale: at 1.00 it predicts 0.700 -> 0.888 on '
     'the first finishing stand, a 27% spread in one pass. A plausible 1.7% spread implies about '
     '0.064.',
     'Because the two factors are on different scales, the 1.00 default you give for the '
     'round-to-flat factor offers no guidance for the flat-to-flat one - and the flat-to-flat '
     'factor does not appear in the legend at all. Note also that the first finishing stand runs '
     '8" rolls against 6" at the second and third, so one value may not serve all three.',
     'Are the three factors single numbers or lookup tables, and if tables, what is the key - '
     'alloy, stand, gauge range?'),

    (7, 'Theoretical & Calculated Width Flat to Flat', 'Calculator will show',
     'One bracket is unclosed as printed, and the two candidate readings differ substantially.',
     'At the first finishing stand, 8" rolls, 0.085 -> 0.045 on a 0.700 entry, factor 1.00:\n'
     '  sqrt(R x (T1 - T2)) = sqrt(4.000 x 0.040) = 0.4000  ->  width 0.8882\n'
     '  sqrt(R) x (T1 - T2) = 2.000 x 0.040       = 0.0800  ->  width 0.7376\n'
     'We have taken the first, which also makes the term identical to your "Contact-Length '
     'Approximation".',
     'A 0.15 in difference in predicted width on one stand. The agreement with your own contact-'
     'length formula is our reason for preferring the first reading, but it is an inference.',
     'Can you confirm the grouping, and that the square root covers roll radius x thickness '
     'change together?'),
]

ws, hr = sheet(wb, 'Points', 'Seven readings we could not settle from the page. '
               'Three need an answer; four we expect your calculator will show.',
               [('#', 5), ('Formula affected', 34), ('', 18), ('What we found', 40),
                ('Worked figures', 46), ('Why it matters', 44), ('Question', 34),
                ('Your answer', 30), ('OK?', 8)])
ws.cell(row=hr, column=3, value='Needs')
r = hr + 1
for n, formula, kind, found, calc, why, q in POINTS:
    put(ws, r, [n, formula, kind, found, calc, why, q, '', ''],
        fills={3: PRIORITY[kind], 8: RESP_FILL, 9: RESP_FILL},
        height=150)
    ws.cell(row=r, column=5).font = MONO_FONT
    r += 1
ws.auto_filter.ref = f'A{hr}:I{r - 1}'

# =============================================================== SHEET 2 =====
FORMULAS = [
    ('Linear Feet Wire/Rod', 'L_F = tw / (pi r² · 12 · rho)', ''),
    ('Linear Feet Round to Round', 'L_F = L1 (Di / Df)²', 'legend defines D1 / D2, formula writes Di / Df'),
    ('Area Reduction Percentage %', 'RA% = (D1² - D2²) / D1²', ''),
    ('Linear Feet Round to Flat', 'L_F = (pi r² L1) / (W2 T2)', 'point 1 - width x thickness'),
    ('Thickness Reduction %', 't% = ((T1 - T2) / T1) · 100', 'see also Engineering Thickness Reduction %'),
    ('Cumulative True Strain', 'SUM e_i = e_t1 + e_t2 + e_t3', 'point 2'),
    ('Individual True Strain', 'e = ln(X2 / X1)', ''),
    ('True Area Reduction Formula', "q' = ln(a1 / a2)", ''),
    ('Wire/Rod Cross-Sectional Area', 'Ae1 = pi × (D1 / 2)²', ''),
    ('Round to Flat Cross-Sectional Area', 'Ae2 = (W2 × T2) - (4r² - pi r²)', 'point 5 - r must be T2/2'),
    ('Engineering Thickness Reduction %', 'eT = ((D1 - T2) / D1) · 100', 'point 6 - two definitions'),
    ('Cross-Sectional Area Reduction %', 'Z = ((Ae - Af) / Ae) × 100', ''),
    ('Material Elongation (Length Multiplier)', 'dL = A1 / A2', 'legend defines Ae / Af'),
    ('Final Cross-Sectional Area Flat Wire', 'a = W2 - (T2² - (pi T2² / 4))', 'point 4 - missing x T2'),
    ('Theoretical Finish Width Round to Flat', 'wT = (pi D1²) / (4 T2)', ''),
    ('Calculated Width Round to Flat',
     'wC = C5 [ (0.7854 D1²/T2)(1 - 15.8 (1 - 2T2/D1)^2.25 (2R/D1)^-0.82) + 0.1426 D1 (2T2/D1) ]',
     'points 3, 5, 6'),
    ('Linear Feet (Flat to Flat)', 'L_F = (L1 W1 T1) / (W2 T2)', 'point 1 - width x thickness'),
    ('Contact-Length Approximation', 'L = sqrt( R (T1 - T2) )', 'nominal roll radius'),
    ('Theoretical & Calculated Width Flat to Flat',
     'TwC = W1 [ 1 + C6 ((sqrt(R(T1-T2)) / W1)((T1-T2)/T1)) ]', 'points 6, 7'),
    ('Edging Theoretical Calculated Thickness', 'tC = T1 [ 1 + k ((W1 - W2) / W1) ]', 'point 6'),
]
ws2, hr2 = sheet(wb, 'Formulas as read',
                 'Every formula in the document is an image, so this is our transcription. '
                 'Please correct anything we have read wrongly - that alone would resolve most of the points.',
                 [('#', 5), ('Your heading', 42), ('As we have read it', 92), ('Raised as', 30)])
r = hr2 + 1
for i, (head, body, note) in enumerate(FORMULAS, start=1):
    put(ws2, r, [i, head, body, note], fills={} if i % 2 else {1: ALT_FILL, 2: ALT_FILL, 3: ALT_FILL, 4: ALT_FILL},
        height=30)
    ws2.cell(row=r, column=3).font = MONO_FONT
    r += 1

# =============================================================== SHEET 3 =====
ws3, hr3 = sheet(wb, 'Worked example',
                 'Live. Change any green input and every figure below recalculates. '
                 'If our reading of a formula is wrong, this is where it will show.',
                 [('', 44), ('Value', 16), ('Excel does', 62)])
# Row numbers are fixed so the formulas below can reference them directly.
# hr3 is 4, so the first body row is 5.
RD, RT, RR, RC = 6, 7, 8, 9              # inputs: diameter, thickness, radius, factor
RW, RRED = 12, 13                        # predicted width, reduction
RAE, RRND, RRECT, RDIFF = 16, 17, 18, 19  # areas
REL_R, REL_S, RFT_R, RFT_S, RFT_D = 22, 23, 24, 25, 26

WIDTH_F = (f'=B{RC}*((0.7854*B{RD}^2/B{RT})*(1-15.8*(1-(2*B{RT}/B{RD}))^2.25'
           f'*((2*B{RR}/B{RD})^-0.82))+0.1426*B{RD}*(2*B{RT}/B{RD}))')

rows = [
    (5, 'INPUTS - change any of these', None, None),
    (RD, 'Entry diameter into the 12 in mill, in   (after the draw boxes)', D1, 'overwrite freely'),
    (RT, 'Target final thickness, in', 0.085, ''),
    (RR, 'Roll radius, in     (12 in mill = 6.000)', 6.000, '8 in stand = 4.000, 6 in = 3.000'),
    (RC, 'Round-to-flat spread factor', 1.000, 'your stated default is 1.00'),
    (11, 'CALCULATED WIDTH ROUND TO FLAT', None, None),
    (RW, 'Predicted width, in', WIDTH_F,
     'shows #NUM! once thickness reaches half the entry diameter - point 3'),
    (RRED, 'Thickness reduction, %', f'=(1-B{RT}/B{RD})*100', ''),
    (15, 'CROSS-SECTIONAL AREA - THE TWO BASES  (point 1)', None, None),
    (RAE, 'Entry area, sq in     (Wire/Rod Cross-Sectional Area)', f'=PI()*B{RD}^2/4', ''),
    (RRND, 'Exit area, rounded edge, sq in', f'=B{RW}*B{RT}-(B{RT}^2-PI()*B{RT}^2/4)',
     'what your two area formulas give'),
    (RRECT, 'Exit area, width x thickness, sq in', f'=B{RW}*B{RT}',
     'what your two linear-feet formulas use'),
    (RDIFF, 'Difference, % of the rectangle', f'=(B{RRECT}-B{RRND})/B{RRECT}*100', ''),
    (21, 'ELONGATION   (Material Elongation, entry area over exit area)', None, None),
    (REL_R, 'on the rounded-edge area', f'=B{RAE}/B{RRND}', 'must be at least 1.000'),
    (REL_S, 'on width x thickness', f'=B{RAE}/B{RRECT}',
     'below 1.000 means the wire got shorter - impossible'),
    (RFT_R, 'Delivered footage per 1,000 ft in - rounded edge', f'=1000*B{REL_R}', ''),
    (RFT_S, 'Delivered footage per 1,000 ft in - width x thickness', f'=1000*B{REL_S}', ''),
    (RFT_D, 'Difference, ft', f'=B{RFT_R}-B{RFT_S}', ''),
]
for r, label, val, note in rows:
    if val is None:
        put(ws3, r, [label, '', ''], font=BOLD_BODY)
        ws3.cell(row=r, column=1).fill = ALT_FILL
        continue
    put(ws3, r, [label, val, note])
    cell = ws3.cell(row=r, column=2)
    if isinstance(val, float):
        cell.fill = INPUT_FILL
        cell.number_format = '0.0000'
    else:
        cell.number_format = '0.000000' if r in (RAE, RRND, RRECT) else '0.000'
ws3.cell(row=REL_S, column=2).fill = RESP_FILL

# =============================================================== SHEET 4 =====
ws4, hr4 = sheet(wb, 'Strain check',
                 'Live, and the point of point 2. The SUM column is a formula, not a typed zero. '
                 'Change any width or thickness and it stays at zero.',
                 [('Stand', 14), ('Rolls', 10), ('Entry T1', 11), ('Entry W1', 11),
                  ('Exit T2', 11), ('Exit W2', 11), ('Length ratio', 13),
                  ('e length', 11), ('e width', 11), ('e thickness', 12), ('SUM', 12)])
chain = [('Finishing 1', '8 in', 0.085, 0.700, 0.045, 0.712),
         ('Finishing 2', '6 in', 0.045, 0.712, 0.025, 0.706),
         ('Finishing 3', '6 in', 0.025, 0.706, 0.016, 0.699)]
r = hr4 + 1
for name, rolls, t1, w1, t2, w2 in chain:
    put(ws4, r, [name, rolls, t1, w1, t2, w2,
                 f'=(D{r}*C{r})/(F{r}*E{r})',
                 f'=LN(G{r})', f'=LN(F{r}/D{r})', f'=LN(E{r}/C{r})',
                 f'=H{r}+I{r}+J{r}'])
    for col in (3, 4, 5, 6):
        ws4.cell(row=r, column=col).fill = INPUT_FILL
        ws4.cell(row=r, column=col).number_format = '0.000'
    for col in (7, 8, 9, 10):
        ws4.cell(row=r, column=col).number_format = '0.0000'
    ws4.cell(row=r, column=11).number_format = '0.000000'
    ws4.cell(row=r, column=11).fill = OK_FILL
    ws4.cell(row=r, column=11).font = BOLD_BODY
    r += 1
r += 1
put(ws4, r, ['The flattening pass is different - it returns the same constant for every product:'],
    font=BOLD_BODY)
r += 2
put(ws4, r, ['Entry D1', 'Exit T2', 'Exit W2', 'Length ratio', 'e length', 'e width',
             'e thickness', 'SUM'], font=BOLD_BODY)
for c in range(1, 9):
    ws4.cell(row=r, column=c).fill = HEAD_FILL
    ws4.cell(row=r, column=c).font = HEAD_FONT
r += 1
for d, t2, w2 in ((D1, 0.085, 0.700), (solve_D1(0.110, 0.625, 6.0), 0.110, 0.625),
                  (solve_D1(0.125, 0.875, 6.0), 0.125, 0.875)):
    put(ws4, r, [round(d, 4), t2, w2,
                 f'=(PI()*A{r}^2/4)/(C{r}*B{r})',
                 f'=LN(D{r})', f'=LN(C{r}/A{r})', f'=LN(B{r}/A{r})',
                 f'=E{r}+F{r}+G{r}'])
    for c in (1, 2, 3):
        ws4.cell(row=r, column=c).fill = INPUT_FILL
        ws4.cell(row=r, column=c).number_format = '0.0000'
    for c in (4, 5, 6, 7):
        ws4.cell(row=r, column=c).number_format = '0.0000'
    ws4.cell(row=r, column=8).number_format = '0.000000'
    ws4.cell(row=r, column=8).fill = BAD_FILL
    ws4.cell(row=r, column=8).font = BOLD_BODY
    r += 1
r += 1
put(ws4, r, ['That constant is ln(pi/4) = -0.241564, the ratio between a round section\'s '
             'width x thickness and its true area. It carries no information about the pass.'],
    font=SUB_FONT)

# ---- leakage guard ----------------------------------------------------------
# Three narrow patterns. The client's own symbols - D1, T2, W2, C5, r - and the
# solidus used for division are legitimate throughout, so neither may be banned
# wholesale; an earlier revision did and flagged twenty false positives.
IDS = re.compile(r'\b(?:FR|TC|OI|FW|TB|DBD|SIM|REQ|ARC|API|PLC|DEP|INT|SEC)-\d{1,3}\b'
                 r'|\bPSG-[A-Z]\d{1,3}\b'
                 r'|\b[GQ]\d{1,3}\b'
                 r'|\bD-\d{1,3}\b'
                 r'|\bV\d{2}\b')
FILES = re.compile(r'\.(?:md|sql|py|xlsx|docx|html|json)\b'
                   r'|\b\d\d-[a-z]+/|\btools/|\bdeliverables/|\bC:\\', re.I)
NAMES = re.compile(r'PassSchedule|FlatWireDB|CoilOutput|ToolingInventory|MasterSpecification'
                   r'|Orchestration|RunReading|SpoolProcessing|FlatWireRun|DieHistory', re.I)
hits = []
for w in wb.worksheets:
    for row in w.iter_rows():
        for c in row:
            if isinstance(c.value, str) and not c.value.startswith('='):
                for line in c.value.split('\n'):
                    for rx, kind in ((IDS, 'identifier'), (FILES, 'file/path'), (NAMES, 'table')):
                        m = rx.search(line)
                        if m:
                            hits.append((w.title, c.coordinate, f'{kind} {m.group(0)!r}', line[:60]))
if hits:
    print('FATAL - client-facing leakage:')
    for h in hits:
        print('   %s!%s  %r  in  %s' % h)
    sys.exit(1)

out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('deliverables/FlatWire_FormulaReview.xlsx')
out.parent.mkdir(parents=True, exist_ok=True)
wb.save(out)
print('wrote %s' % out)
print('  Points           : %d rows' % len(POINTS))
print('  Formulas as read : %d rows' % len(FORMULAS))
print('  Worked example   : live, %d inputs' % 4)
print('  Strain check     : live, %d passes' % (len(chain) + 3))
print('  leakage scan     : clean')
