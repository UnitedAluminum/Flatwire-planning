"""
Build the six-screen trial run workbook.

    python MVP-1/ProjectPlan/Tools/build_trial_run_xlsx.py [output.xlsx]

With no arguments it writes MVP-1/ProjectPlan/Development/FlatWire_TrialRunPlan.xlsx.

WHY THIS EXISTS
    TrialRunPlan.md is the plan of record for the trial run, and it is a long document. The
    people who act on it - delivery leads assigning work, and the client tracking what is being
    built - need it as a sheet they can filter and sort. Effort figures in this repository
    propagate to about twenty files, so a hand-maintained second copy of the trial numbers is
    the most likely way they drift. Everything numeric here is therefore PARSED from the plan,
    and a reconciliation guard makes a silent parse failure impossible too.

WHERE EACH FIELD COMES FROM
    Structure  - every figure, story identifier, sprint, date, phase and blocker - is parsed
                 from ../Development/TrialRunPlan.md. Change the plan and re-run.
    Titles     - parsed from ../Development/TaskBreakdown.md, which holds all 117 story bodies.
                 It is the title source rather than StaffedSprintPlans.md because that document
                 predates FW-202/203/204 and does not carry them.
    Prose      - the plain-language "what it delivers" per item, the Read Me and the blocker
                 phrasing - comes from TrialRunContent.md, the only place it is authored.
    "Reference title" is read from the content file for the drift guard alone and is never
    written to the workbook.

THREE DELIBERATE DEPARTURES FROM build_development_plan_xlsx.py
    Do not "fix" these back. Each is a decision recorded on 14 Aug 2026.

    1. STORY IDENTIFIERS ARE PRESENT, AND THERE IS NO LEAKAGE GUARD.
       That script lists (r'\\bFW-\\d', 'backlog identifier') in a fatal LEAKS list and numbers
       its rows 1..N instead. This workbook exists so the client can refer to a task by its
       identifier, which is the opposite requirement. It is INTERNAL, shared with the client,
       and lives in Development/ rather than MVP-1/SRS/ for exactly that reason - MVP-1/SRS/ is
       where leakage-guarded client deliverables live.
    2. EFFORT IS IN HOURS, NOT DAYS.
       That script presents days only, on the grounds that hours invite a rate conversation.
       TrialRunPlan.md is in hours and the audience is internal, so hours are the primary unit.
       A Days column is derived at HOURS_PER_DAY as a reading aid. That constant is 6.5 since
       25 Aug 2026, so a "day" in this workbook is a 6.5 h day and the column header says so.
       Do not read these as 8 h days - the header is the only thing stopping that.
    3. AN ABBREVIATION GUARD TAKES THE LEAKAGE GUARD'S PLACE.
       No bare abbreviation may reach a cell - see LABELS and check_abbreviations below. The
       plan writes FE/BE/DB/RT/T1/1A/DB5/FL2 freely because it is read as prose; a filtered
       spreadsheet column has no surrounding sentence, and DB is actively ambiguous - it is
       both the Database discipline and the Dashboard prefix.

FIVE GUARDS, ALL FATAL (plus one warning)
    1. Reconciliation - every total in the plan is recomputed from its own cells: the block and
                        discipline tables and the sprint tables all sum to 832; the phase-by-
                        discipline grid agrees on every row, column and corner; the three
                        sub-phases sum to 462; the all-in table's columns equal its total row
                        AND its development column equals those same 462; the deferred table
                        sums to 330.
                        EXTENDED 15 AUG 2026, and the reason is worth keeping: the 778 -> 806
                        re-baseline reached sections 1.3 and 4, all five guards passed, and
                        section 2 - capacity, utilisation, staffing options, margin - was left
                        a whole revision behind, because nothing here read section 2. It now
                        checks section 2's effort figure and every sprint's planned hours
                        against section 4's allocation, and it reads the hand-coded baseline
                        off the backlog's own reconciliation lines instead of trusting the
                        plan's transcription of them.
                        A sprint planned above its capacity is a WARNING, not a failure - see
                        guard()'s docstring.
    2. Coverage       - every work item parsed from the plan has a content entry and vice versa.
    3. Title          - every identifier in the plan resolves to a title in TaskBreakdown.md.
                        An unresolved identifier is fatal, never blank-filled.
    4. Drift          - the content entry's recorded reference title still matches the plan's.
                        Catches a renumbering silently moving prose onto the wrong work item.
    5. Abbreviation   - every cell of the built workbook is re-read from disk and scanned. The
                        Read Me glossary is the only exempt sheet.
    On any failure the workbook is deleted and the build exits non-zero, so a file on disk
    means all five passed.

REQUIREMENTS
    openpyxl (tested against 3.1.5). Nothing else.

GOTCHA
    If the output is open in Excel the write fails with PermissionError - a ~$ lock file in
    Development/ is the tell. Close it and re-run.
"""
import os, re, sys

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

HERE = os.path.dirname(os.path.abspath(__file__))
def _repo_root():
    """Walk up to the directory holding .git.

    Deliberately NOT a level count. These scripts have moved twice and a hard-coded
    depth broke them both times; a marker survives the next move too.
    """
    d = os.path.abspath(os.path.dirname(__file__))
    while d != os.path.dirname(d):
        if os.path.isdir(os.path.join(d, '.git')):
            return d
        d = os.path.dirname(d)
    raise SystemExit('cannot locate repository root (no .git found above %s)' % __file__)


ROOT = _repo_root()
DEV = os.path.join(ROOT, '60-delivery')

PLAN = os.path.join(DEV, 'TrialRunPlan.md')
TASKS = os.path.join(DEV, 'TaskBreakdown.md')
CONTENT = os.path.join(HERE, 'TrialRunContent.md')
DEFAULT_OUT = os.path.join(DEV, 'FlatWire_TrialRunPlan.xlsx')

ISSUE_DATE = 'August 25, 2026'
HOURS_PER_DAY = 6.5        # 8 -> 6.5 on 25 Aug 2026 with the three-resource re-baseline.
                           # This is an AVAILABILITY figure, not a productivity one: each resource
                           # commits 6.5 h of the working day to flat wire. The estimates stay
                           # calendar-hour estimates and TOTAL_HOURS does not move - see the note
                           # on the Days column below, and TrialRunPlan.md section 2.
TOTAL_HOURS = 869          # the plan's own headline; every guard reconciles to it
                           # 829 -> 869 on 18 Aug 2026: FW-219, the FL2/FL3 run-end write-back
                           # into the shared schema (40 h, T3). See TrialRunPlan.md section 5.3.
                           # 832 -> 829 on 18 Aug 2026: D-32 cancels the shared-schema
                           # migration, so FW-002's 3 h leaves 1C's trial scope.
PLATFORM_HOURS = 459       # Phase 1A + 1B + 1C  (462 -> 459 with the same change)
DEFERRED_HOURS = 330

GLOSSARY_SHEET = 'Read Me'


# --------------------------------------------------------------- label expansion
# Applied to EVERY cell before write, parsed figures included.
#
# Two rules make this safe, and both matter:
#   1. SINGLE PASS. All keys go into one alternation, longest first, and each position in the
#      string is consumed once. Looping the keys instead would let one replacement's output be
#      re-matched by a later key.
#   2. NO REPLACEMENT CONTAINS ITS OWN KEY. That is what makes expand() idempotent, so a cell
#      that has already been through it is left alone. The phrase entries at the top exist for
#      exactly this reason - "1A" alone cannot expand to "Phase 1A - ..." without re-matching.
LABELS = {
    # Phrases first - they remove several tokens at once and keep the sentence readable.
    '(1A Angular · 1B Backend · 1C Database)': '(Angular, Backend and Database Foundations)',
    'Phase 1A ∥ 1B ∥ 1C': 'Angular, Backend and Database Foundations, in parallel',
    '1B close-out': 'Backend foundation close-out',
    'DI-swappable': 'Dependency-injection-swappable',
    'DI registration': 'Dependency-injection registration',
    'Dapper/EF': 'Dapper and Entity Framework',
    'PWA cache sync': 'Offline cache sync',
    'JWT authentication': 'Token-based authentication',
    'OPC ingest': 'Plant data ingest',
    'OPC feed simulator': 'Plant data feed simulator',
    'no DB2A staging': 'no rod pre-staging',
    'cross-DB': 'cross-database',
    'DB3-FL2': 'the Active Run Monitor for Flattening Line 2',
    'FL2/FL3': 'Flattening Lines 2 and 3',
    # Screens.
    'DB2A': 'Rod Pre-Check-in station',
    'DB10': 'Supervisor Shift Summary screen',
    'DB11': 'Roll Adjust dialog',
    'DB12': 'Rod Checkout dialog',
    'DB1': 'Line Status Overview screen',
    'DB2': 'Rod Check-in screen',
    'DB3': 'Active Run Monitor',
    'DB5A': 'Spool Queue screen',
    'DB5': 'Spool Check-in screen',
    'DB6': 'Quality Checkpoint dialog',
    'DB7b': 'Packing Station screen',
    'DB7': 'Output Coil Completion screen',
    'DB8': 'Material Rejection dialog',
    'DB9A': 'Pass Schedule List screen',
    'DB9': 'Pass Schedule Management screen',
    # Equipment and plant vocabulary.
    'FL1': 'Flattening Line 1',
    'FL2': 'Flattening Line 2',
    'FL3': 'Flattening Line 3 (hybrid)',
    'FM1': 'Finishing Mill 1',
    'FM2': 'Finishing Mill 2',
    'TPO': 'Traversing Payoff',
    'SPC': 'Statistical Process Control',
    'WIP': 'Work-in-Progress',
    'UAT': 'User Acceptance Testing',
    'FTE': 'full-time equivalent',
    'DDL': 'database build script',
    'PLC': 'Programmable Logic Controller',
    'OPC': 'plant data',
    # Sprints. Short form inline; the Sprint Plan sheet writes the dated form explicitly.
    'T1': 'Sprint 1', 'T2': 'Sprint 2', 'T3': 'Sprint 3', 'T4': 'Sprint 4',
    'S0': 'gate sprint', 'S1': 'sprint one', 'S2': 'sprint two', 'S3': 'sprint three',
    # Disciplines.
    'FE': 'Frontend (Angular)',
    'BE': 'Backend (.NET)',
    'DB': 'Database (SQL Server)',
    'RT': 'Real-time and controller integration',
    'QA': 'Quality Assurance',
    'BA': 'Business Analysis',
}

# One alternation, longest key first, bare occurrences only.
#
# The boundary is \w rather than [\w-], which MUST match the guard's \b or the two disagree:
# a hyphen-joined compound like "DB3-FL2" is two abbreviations, not one identifier, so it has
# to expand - and it was flagged-but-unexpanded while these two rules differed. Code
# identifiers stay intact because they join on word characters or underscores, never hyphens:
# PLCTagService, FlatWireDbContext and sp_GetGaugeTrace are all safe. Hyphenated component
# names are safe too - no segment of run-status-cards or chart-tab-strip is an abbreviation.
_LABEL_RE = re.compile(
    r'(?<!\w)(' + '|'.join(re.escape(k) for k in
                           sorted(LABELS, key=len, reverse=True)) + r')(?!\w)')

DISCIPLINE = {'FE': 'Frontend (Angular)', 'BE': 'Backend (.NET)',
              'DB': 'Database (SQL Server)', 'RT': 'Real-time and controller integration'}
SUBPHASE = {'1A': 'Phase 1A — Angular Foundation',
            '1B': 'Phase 1B — Backend Foundation',
            '1C': 'Phase 1C — Database Foundation'}

# Bare tokens that must never survive into a cell. DB and DB\d are separate branches because
# the first is the discipline and the second the screen prefix - the ambiguity this guard
# exists for.
#
# THIS SET IS DELIBERATELY WIDER THAN LABELS, and the asymmetry is the point. LABELS expands
# what has a good plain-language equivalent; anything left over is caught here and has to be
# fixed at the source. EF, DI, PWA and JWT appear below but only inside PHRASES in LABELS
# ("Dapper/EF", "JWT authentication"), so a bare one in newly authored prose is flagged rather
# than silently expanded to something clumsy. Verified by injecting a bare "JWT" and "DI" into
# a content entry: both were named with sheet and cell, and no workbook was left on disk.
#
# Two deliberate exemptions:
#   * "Phase 1A" and "Sprint 1" are IDENTIFIERS, not shorthand, and read perfectly. Only a
#     BARE "1A" is flagged.
#   * Code identifiers are not caught at all: \bPLC\b does not match PLCTagService, because
#     \b needs a non-word character and T is one.
ABBREV = re.compile(
    r'(?<!Phase )\b(?:FE|BE|DB|RT|QA|BA|T[1-4]|S[0-3]|1[ABC]'
    r'|DB\d{1,2}[A-Za-z]?|FL[123]|FM[12]|TPO|SPC|WIP|UAT|FTE|DDL|PLC|OPC|EF|DI|PWA|JWT)\b')


def expand(text):
    """Write every abbreviation out, in one pass. Idempotent."""
    if not isinstance(text, str):
        return text
    return _LABEL_RE.sub(lambda m: LABELS[m.group(1)], text)


def phase_label(value):
    """'1A Angular foundation' -> 'Phase 1A — Angular Foundation'.

    Structural, not a substitution: the whole cell is replaced, so nothing is left to
    re-expand and the duplicated trailing words are dropped.
    """
    m = re.match(r'^(1[ABC])\b', value.strip())
    return SUBPHASE[m.group(1)] if m else value


# ----------------------------------------------------------------- markdown helpers
# clean/render/_fold_blank_runs/norm are the two existing workbook builders' helpers,
# unchanged. Copied rather than imported - see Tools/README.md: each builder must keep
# building if another is edited. The rich-text fold is a real openpyxl 3.1.5 bug fix.

def clean(text):
    """Strip the markdown a table cell carries."""
    text = re.sub(r'\[([^\]]*)\]\([^)]*\)', r'\1', text)   # links -> label
    text = text.replace('`', '').replace('**', '').replace('*', '')
    return re.sub(r'\s+', ' ', text).strip()


_BOLD = InlineFont(rFont='Calibri', sz=11, b=True)
_ITALIC = InlineFont(rFont='Calibri', sz=11, i=True)
_EMPHASIS = re.compile(r'\*\*(.+?)\*\*|(?<![\w*])\*([^*\n]+?)\*(?![\w*])')


def _fold_blank_runs(parts):
    """Never emit a whitespace-only run.

    openpyxl 3.1.5 writes xml:space="preserve" on a whitespace-EDGED run (' x ') but omits
    it on a whitespace-ONLY one (' '), so XML normalisation strips the content and Excel
    reports 'Repaired Records: String properties' on open. Two bold spans separated by a
    single space produce exactly that run.
    """
    def attach(part, ws, before=False):
        if isinstance(part, str):
            return (ws + part) if before else (part + ws)
        return TextBlock(part.font, (ws + part.text) if before else (part.text + ws))

    out = []
    for part in parts:
        text = part if isinstance(part, str) else part.text
        if text.strip():
            out.append(part)
            continue
        if out:
            out[-1] = attach(out[-1], text)
        elif len(parts) > 1:
            nxt = parts[parts.index(part) + 1]
            parts[parts.index(part) + 1] = attach(nxt, text, before=True)
    return out


def render(text):
    """Markdown emphasis -> Excel rich text, so bold survives into the cell."""
    if not isinstance(text, str):
        return text
    text = text.replace('`', '')
    if '**' not in text and '*' not in text:
        return text
    parts, last = [], 0
    for m in _EMPHASIS.finditer(text):
        if m.start() > last:
            parts.append(text[last:m.start()])
        body = m.group(1) or m.group(2)
        parts.append(TextBlock(_BOLD if m.group(1) else _ITALIC, body))
        last = m.end()
    if last < len(text):
        parts.append(text[last:])
    parts = _fold_blank_runs(parts)
    if not parts:
        return ''
    if len(parts) == 1 and isinstance(parts[0], str):
        return parts[0]
    return CellRichText(*parts)


def norm(text):
    """Normalise a title for the drift guard: letters, digits and single spaces only."""
    text = clean(text).lower()
    text = re.sub(r'[^a-z0-9 ]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def days(hours, dp=1):
    """Hours -> days. A reading aid only; hours are this workbook's unit."""
    return round(hours / HOURS_PER_DAY, dp)


# ----------------------------------------------------------------------- parsers

def _tables_after(text, heading):
    """Every markdown table between a heading and the next heading of the same or higher level.

    Located by heading TEXT rather than section number, so renumbering the plan - which has
    happened twice already - does not silently break the build. Returns a list of
    (header, rows), each cell cleaned of markdown.
    """
    i = text.find(heading)
    if i < 0:
        sys.exit(f'FATAL: heading not found in the plan: {heading!r}')
    depth = len(heading) - len(heading.lstrip('#')) or 6
    tables, header, rows = [], None, []

    def flush():
        if header is not None and rows:
            tables.append((header, list(rows)))

    for line in text[i + len(heading):].splitlines():
        s = line.strip()
        if s.startswith('#') and (len(s) - len(s.lstrip('#'))) <= depth:
            break
        if s.startswith('|'):
            cells = [clean(c) for c in s.strip('|').split('|')]
            if set(''.join(cells)) <= set('-: '):
                continue
            if header is None:
                header = cells
            else:
                rows.append(cells)
        elif header is not None and rows:
            flush()
            header, rows = None, []
        elif header is not None and not rows:
            header = None
    flush()
    return tables


def _rows_after(text, heading, expect=None, nth=1):
    """The nth table under a heading.

    `expect` is a column heading that must appear SOMEWHERE in the header row - cheap
    insurance that an edit to the plan has not silently shifted the tables around. It is
    matched against the whole row rather than the first cell because several of the plan's
    tables deliberately leave the first column heading blank.
    """
    tables = _tables_after(text, heading)
    if len(tables) < nth:
        sys.exit(f'FATAL: expected at least {nth} table(s) under {heading!r}, '
                 f'found {len(tables)}')
    header, rows = tables[nth - 1]
    if expect and not any(expect.lower() in h.lower() for h in header):
        sys.exit(f'FATAL: table {nth} under {heading!r} has columns {header!r}, '
                 f'expected one of them to contain {expect!r}')
    if not rows:
        sys.exit(f'FATAL: no rows in table {nth} under {heading!r}')
    return header, rows


def _int(value, default=None):
    m = re.search(r'-?\d[\d,]*', str(value).replace('−', '-'))
    if not m:
        if default is not None:
            return default
        sys.exit(f'FATAL: expected a number, got {value!r}')
    return int(m.group(0).replace(',', ''))


def parse_handcoded(path):
    """Phase 1A/1B/1C hand-coded all-in totals, read off the backlog's own reconciliation
    lines rather than transcribed into the plan by hand.

    Added 15 Aug 2026. 1B moved 442 -> 541 -> 519 and 1C 215 -> 221 while the plan still
    printed the older pair, so its "43 % reduction against 1,027 h" headline was comparing
    against a baseline that no longer existed. Transcribed figures go stale silently; parsed
    ones cannot.
    """
    text = open(path, encoding='utf-8').read()
    out = {}
    for m in re.finditer(r'\*\*Phase (1[ABC]) reconciliation\*\*(.*)', text):
        # The LAST bolded "N h" on the line is the all-in total; the earlier bolded figures
        # are the base and (for 1B) a deliberately held QA number.
        totals = re.findall(r'\*\*([\d,]+)\s*h\*\*', m.group(2))
        if totals:
            out[m.group(1)] = int(totals[-1].replace(',', ''))
    return out


def parse_titles(path):
    """Story identifier -> its verbatim title, from the backlog's story bodies."""
    text = open(path, encoding='utf-8').read()
    out = {sid: ' '.join(title.split())
           for sid, title in re.findall(r'^###### (FW-[\w\d]+) · (.+?)$', text, re.M)}
    if not out:
        sys.exit(f'FATAL: no story titles parsed from {os.path.relpath(path, ROOT)}')
    return out


def parse_plan(path):
    """Everything numeric, from the plan of record."""
    text = open(path, encoding='utf-8').read()
    p = {}

    # ---- blocks and disciplines (the two tables under 1.3 Total)
    # 1.3 carries two tables before its first sub-heading: blocks, then disciplines.
    # Drop this table's footer - its first cell is blank and it restates the grand total, which
    # would be double-counted. The workbook writes its own Total row instead.
    _, rows = _rows_after(text, '### 1.3 Total', 'Block', nth=1)
    p['blocks'] = [dict(block=r[0], hours=_int(r[1]), share=r[2])
                   for r in rows if r[0].strip() and r[1].strip()]
    _, rows = _rows_after(text, '### 1.3 Total', 'Stream', nth=2)
    p['disciplines'] = [dict(code=r[0].split()[0], name=r[0], hours=_int(r[1]), share=r[2])
                        for r in rows]

    # ---- the phase-by-discipline grid
    _, rows = _rows_after(text, '#### Phase × stream — the staffing grid', 'Phase')
    p['grid'] = [dict(phase=r[0], fe=_int(r[1], 0), be=_int(r[2], 0), db=_int(r[3], 0),
                      rt=_int(r[4], 0), total=_int(r[5]), sprint=r[6])
                 for r in rows if r[5].strip()]

    # ---- Phase 1: three bases, the all-in view, and the three deliverable tables
    _, rows = _rows_after(text, '### 1.4 Phase 1 in detail', 'Hand-coded')
    # Keep the footer row - its first cell is blank by design and it carries the totals every
    # reconciliation guard checks against. Filtering it out is what made 1C look like the total.
    p['bases'] = [dict(subphase=r[0], handcoded=_int(r[1]), ai=_int(r[2]), trial=_int(r[3]),
                       reduction=r[4], t1=_int(r[5], 0), t2=_int(r[6], 0))
                  for r in rows if _int(r[3], 0)]
    _, rows = _rows_after(text, '#### Phase 1 all-in, on the trial\'s own basis', 'Dev')
    p['allin'] = [dict(subphase=r[0], dev=_int(r[1]), qa=_int(r[2]), cont=_int(r[3]),
                       total=_int(r[4])) for r in rows]
    p['platform'] = {}
    for code, heading in (('1A', '#### 1A Angular foundation'),
                          ('1B', '#### 1B Backend foundation'),
                          ('1C', '#### 1C Database foundation')):
        _, rows = _rows_after(text, heading, 'Deliverable')
        items = []
        for r in rows:
            if not r[0].strip() or r[0].startswith('|'):
                continue
            sid = re.search(r'FW-[\w\d]+', r[2]) if len(r) > 2 else None
            items.append(dict(deliverable=r[0], handcoded=_int(r[1], 0),
                              story=sid.group(0) if sid else '',
                              trial=_int(r[-1], 0)))
        p['platform'][code] = [i for i in items if i['story']]

    # ---- the capacity position
    # Read since 15 Aug 2026. NOTHING here read section 2 before that date, which is how the
    # 778 -> 806 re-baseline reached 1.3 and 4 and left 2's effort, utilisation, staffing and
    # margin figures a whole revision behind. A guard that reconciles every total except the
    # one the client reads first is not a guard.
    _, rows = _rows_after(text, '## 2. The capacity position', 'Measure')
    p['capacity'] = {r[0]: r[1] for r in rows if r[0].strip()}

    # ---- the sprint calendar and the staffing options
    _, rows = _rows_after(text, '### 2.1', 'Sprint')
    # The final sprint is testing and sign-off, so it carries no development hours and its
    # capacity, planned and utilisation cells are dashes. Default them to 0 and render blank -
    # a deliberate dash is not a parse failure.
    p['sprints'] = [dict(sprint=r[0], dates=r[1], workdays=_int(r[2], 0), team=r[3],
                         capacity=_int(r[4], 0), planned=_int(r[5], 0),
                         util=r[6], content=r[7])
                    for r in rows if re.fullmatch(r'T[1-4]', r[0])]
    _, rows = _rows_after(text, '### 2.2 What each staffing option lands', 'Option')
    p['options'] = [dict(option=r[0], complete=r[1], signoff=r[2], note=r[3]) for r in rows]
    _, rows = _rows_after(text, '### 2.3 Reserves', 'Reserve')
    p['reserves'] = [dict(reserve=r[0], hours=r[1], basis=r[2]) for r in rows]

    # ---- the allocation: every work item, its hours and its sprint
    sec = text.split('## 4. Story → sprint allocation')[1].split('### Deferred')[0]
    p['items'], p['sprint_totals'] = [], {}
    for code, heading in (('T1', '### T1 —'), ('T2', '### T2 —'), ('T3', '### T3 —')):
        head, block = sec.split(heading)[1].split('\n', 1)[0], sec.split(heading)[1]
        block = block.split('\n###')[0]
        p['sprint_totals'][code] = _int(re.search(r'·\s*([\d,]+)\s*h', head).group(1))
        by_row = {}
        for line in block.splitlines():
            if not line.strip().startswith('|'):
                continue
            cells = [c.strip() for c in line.strip().strip('|').split('|')]
            if len(cells) < 3 or set(''.join(cells)) <= set('-: '):
                continue
            group = clean(cells[0])
            for sid, hrs in re.findall(r'`(FW-[A-Z0-9\-]+)`\*{0,2}\s*(\d+)', cells[1]):
                by_row[sid] = (group, int(hrs))
        for sid, (group, hrs) in by_row.items():
            p['items'].append(dict(story=sid, sprint=code, group=group, hours=hrs))

    # ---- deferred
    _, rows = _rows_after(text, '### Deferred — still MVP-1', 'Deferred')
    p['deferred'] = [dict(item=r[0], hours=_int(r[1]), reason=r[2]) for r in rows]

    # ---- blockers
    _, rows = _rows_after(text, '## 6. Blockers', '#')
    p['blockers'] = [dict(n=r[0], blocker=r[1], blocks=r[2], needed=r[3]) for r in rows]

    # ---- the two removals
    # Each removal sub-section carries a Consequence/Handling table. The pre-check-in one is
    # preceded by a "blockers that leave with it" table, so it is the SECOND table there.
    p['removals'] = []
    for name, heading, nth in (
            ('Dashboard 1 — Line Status Overview', '#### DB1 Line Status Overview', 1),
            ('Dashboard 2A — Rod Pre-Check-in, and weld capture',
             '#### DB2A Pre-Check-in + weld capture', 2)):
        _, rows = _rows_after(text, heading, 'Consequence', nth=nth)
        for r in rows:
            p['removals'].append(dict(removal=name, consequence=r[0], handling=r[1]))

    # The blockers that left trial scope with the pre-check-in removal - worth carrying,
    # because "this removal took two blockers with it" is the strongest thing about it.
    _, rows = _rows_after(text, '#### DB2A Pre-Check-in + weld capture', 'Blocker', nth=1)
    p['removal_upside'] = [dict(blocker=r[0], why=r[1]) for r in rows]
    return p


def parse_content(path):
    """Authored prose, keyed by identifier. One section per part."""
    text = open(path, encoding='utf-8').read()
    parts, current = {}, None
    for block in re.split(r'^## ', text, flags=re.M)[1:]:
        name = block.splitlines()[0].strip()
        if name in ('Read Me', 'Glossary', 'Work Items', 'Blockers'):
            current = name
            parts[current] = {}
            body = block
        else:
            continue
        for entry in re.split(r'^### ', body, flags=re.M)[1:]:
            key = entry.splitlines()[0].strip()
            fields = dict(re.findall(r'^\*\*(.+?):\*\*\s*(.+?)$', entry, re.M))
            parts[current][key] = fields
        if current == 'Read Me':
            parts[current] = dict(re.findall(r'^\*\*(.+?):\*\*\s*(.+?)$', body, re.M))
        if current == 'Glossary':
            _, rows = _rows_after(body, 'Glossary', 'Short form')
            parts[current] = [(r[0], r[1]) for r in rows]
    for need in ('Read Me', 'Glossary', 'Work Items', 'Blockers'):
        if not parts.get(need):
            sys.exit(f'FATAL: content file has no usable "{need}" section')
    return parts


# ------------------------------------------------------------------------- guards

def guard(plan, titles, content, handcoded):
    """Reconciliation, coverage, title resolution and drift. Every failure is collected so
    one run reports all of them rather than the first.

    One thing here is a WARNING and not a failure, deliberately: a sprint planned above its
    own capacity. That is a true statement about the plan, not a defect in it - section 2.1
    documents T1 at 409 h against 400 h and names three ways out. A guard that refused to
    build a document for saying something uncomfortable would just get switched off.
    """
    bad = []

    def check(label, got, want):
        if got != want:
            bad.append(f'{label}: computed {got}, the plan prints {want}')

    # 1 - reconciliation
    check('block table', sum(b['hours'] for b in plan['blocks']), TOTAL_HOURS)
    check('discipline table', sum(d['hours'] for d in plan['disciplines']), TOTAL_HOURS)
    check('sprint totals', sum(plan['sprint_totals'].values()), TOTAL_HOURS)
    check('allocated work items', sum(i['hours'] for i in plan['items']), TOTAL_HOURS)
    for code, total in plan['sprint_totals'].items():
        check(f'sprint {code} items',
              sum(i['hours'] for i in plan['items'] if i['sprint'] == code), total)
    for row in plan['grid']:
        check(f"grid row {row['phase']}",
              row['fe'] + row['be'] + row['db'] + row['rt'], row['total'])
    body, footer = plan['grid'][:-1], plan['grid'][-1]
    for key in ('fe', 'be', 'db', 'rt', 'total'):
        check(f'grid column {key}', sum(r[key] for r in body), footer[key])
    bases, base_total = plan['bases'][:-1], plan['bases'][-1]
    check('phase 1 trial', sum(b['trial'] for b in bases), PLATFORM_HOURS)
    check('phase 1 trial footer', base_total['trial'], PLATFORM_HOURS)
    for key in ('handcoded', 'ai', 'trial', 't1', 't2'):
        check(f'three-bases column {key}', sum(b[key] for b in bases), base_total[key])
    allin, allin_total = plan['allin'][:-1], plan['allin'][-1]
    for row in plan['allin']:
        check(f"all-in row {row['subphase']}", row['dev'] + row['qa'] + row['cont'],
              row['total'])
    for key in ('dev', 'qa', 'cont', 'total'):
        check(f'all-in column {key}', sum(r[key] for r in allin), allin_total[key])
    for code, items in plan['platform'].items():
        want = next(b['trial'] for b in bases if b['subphase'].startswith(code))
        check(f'sub-phase {code} deliverables', sum(i['trial'] for i in items), want)
    check('deferred table', sum(d['hours'] for d in plan['deferred']), DEFERRED_HOURS)

    # 1b - section 2, the capacity position. ADDED 15 AUG 2026, and this is the gap that let
    # a whole stale revision ship: the 778 -> 806 re-baseline reached 1.3 and 4, every guard
    # above passed, and section 2 - the first thing a reader looks at - still said 778 h,
    # 101 % utilisation and 70 h of margin. Nothing read it.
    effort = next((v for k, v in plan['capacity'].items()
                   if 'development effort' in k.lower()), None)
    if effort is None:
        bad.append('section 2 has no "Trial development effort" row to reconcile')
    else:
        check('section 2 development effort', _int(effort), TOTAL_HOURS)
    by_code = {s['sprint']: s for s in plan['sprints']}
    for code, total in plan['sprint_totals'].items():
        s = by_code.get(code)
        if s is None:
            bad.append(f'sprint {code} is allocated in section 4 but absent from 2.1')
            continue
        check(f'sprint {code} planned in 2.1', s['planned'], total)
    # The all-in table is derived from the same sub-phases as the three-bases table and had
    # drifted from it by 28 h - it was still carrying a 1B of 192.
    check('all-in dev column', sum(r['dev'] for r in allin), PLATFORM_HOURS)
    # The hand-coded baseline is the backlog's, not ours to transcribe.
    for b in bases:
        code = b['subphase'].strip('*').split()[0]
        if code in handcoded:
            check(f'{code} hand-coded against the backlog', b['handcoded'], handcoded[code])

    # 2 - coverage, both directions
    planned = {i['story'] for i in plan['items']}
    authored = set(content['Work Items'])
    for sid in sorted(planned - authored):
        bad.append(f'{sid} is in the plan with no content entry')
    for sid in sorted(authored - planned):
        bad.append(f'{sid} has a content entry but is not in the plan')

    # 3 - title resolution
    for sid in sorted(planned):
        if sid not in titles:
            bad.append(f'{sid} has no title in the backlog - never blank-fill an identifier')

    # 4 - drift
    for sid in sorted(planned & authored):
        entry = content['Work Items'][sid]
        for field in ('Reference title', 'Work item', 'Delivers'):
            if not entry.get(field, '').strip():
                bad.append(f'{sid} content entry is missing "{field}"')
        ref, real = entry.get('Reference title', ''), titles.get(sid, '')
        if ref and real and norm(ref) != norm(real):
            bad.append(f'{sid} drift: content says {clean(ref)!r}, '
                       f'the backlog says {clean(real)!r}')

    # blockers
    if len(content['Blockers']) < len(plan['blockers']):
        bad.append(f"{len(plan['blockers'])} blockers in the plan, "
                   f"{len(content['Blockers'])} authored")
    return bad


def check_abbreviations(path):
    """Re-read every cell from disk and refuse any bare abbreviation. The glossary is exempt."""
    wb = load_workbook(path, rich_text=True)
    hits = []
    for ws in wb.worksheets:
        if ws.title == GLOSSARY_SHEET:
            continue
        for row in ws.iter_rows():
            for cell in row:
                text = str(cell.value) if cell.value is not None else ''
                for m in set(ABBREV.findall(text)):
                    hits.append(f'{ws.title}!{cell.coordinate}: bare "{m}" in {text[:70]!r}')
    return hits


# --------------------------------------------------------------------- rendering
HEAD_FILL = PatternFill('solid', fgColor='1F3864')
TITLE_FONT = Font(name='Calibri', size=15, bold=True, color='1F3864')
SUB_FONT = Font(name='Calibri', size=10, italic=True, color='404040')
SECTION_FONT = Font(name='Calibri', size=12, bold=True, color='1F3864')
HEAD_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
BODY_FONT = Font(name='Calibri', size=11)
ALT_FILL = PatternFill('solid', fgColor='F5F7FA')
FOCUS_FILL = PatternFill('solid', fgColor='FFF7E0')
TOTAL_FILL = PatternFill('solid', fgColor='DCE3EF')
TOTAL_FONT = Font(name='Calibri', size=11, bold=True)
THIN = Side(style='thin', color='C8CDD3')
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
HEAD_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
CENTRE = Alignment(horizontal='center', vertical='top')


def sheet_header(ws, title, subtitle, tab=None):
    ws.sheet_view.showGridLines = False
    if tab:
        ws.sheet_properties.tabColor = tab
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 22
    ws['A2'] = subtitle
    ws['A2'].font = SUB_FONT
    ws.row_dimensions[2].height = 14


def widths(ws, columns):
    for i, (_, w) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def table(ws, start_row, columns, rows, centre_cols=(), zebra=True, focus_cols=(),
          total_rows=()):
    """Header row plus data at start_row. Returns the row after the table."""
    for i, (head, _) in enumerate(columns, start=1):
        c = ws.cell(row=start_row, column=i, value=expand(head))
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = HEAD_ALIGN
        c.border = CELL_BORDER
    ws.row_dimensions[start_row].height = 30

    for r, values in enumerate(rows, start=start_row + 1):
        is_total = (r - start_row - 1) in total_rows
        for i, value in enumerate(values, start=1):
            value = expand(value) if isinstance(value, str) else value
            c = ws.cell(row=r, column=i,
                        value=render(value) if isinstance(value, str) and value else value)
            c.font = TOTAL_FONT if is_total else BODY_FONT
            c.alignment = CENTRE if (i - 1) in centre_cols else TOP_WRAP
            c.border = CELL_BORDER
            if is_total:
                c.fill = TOTAL_FILL
            elif (i - 1) in focus_cols:
                c.fill = FOCUS_FILL
            elif zebra and (r - start_row) % 2 == 0:
                c.fill = ALT_FILL
    return start_row + len(rows) + 1


def section(ws, row, text):
    c = ws.cell(row=row, column=1, value=expand(text))
    c.font = SECTION_FONT
    return row + 1


def note(ws, row, lines):
    for line in lines:
        c = ws.cell(row=row, column=1, value=render(expand(line)))
        c.font = SUB_FONT
        c.alignment = Alignment(vertical='top', wrap_text=True)
        row += 1
    return row


# ------------------------------------------------------------------------- build

def build(out_path):
    titles = parse_titles(TASKS)
    plan = parse_plan(PLAN)
    content = parse_content(CONTENT)

    failures = guard(plan, titles, content, parse_handcoded(TASKS))
    if failures:
        print('FATAL - build refused:')
        for f in failures:
            print(f'  * {f}')
        if os.path.exists(out_path):
            os.remove(out_path)
        sys.exit(1)

    # Not fatal - see guard()'s docstring. Printed after the failures so it is the last thing
    # on screen on a successful build rather than the first thing scrolled past.
    over = [s for s in plan['sprints'] if s['capacity'] and s['planned'] > s['capacity']]
    for s in over:
        print(f'  ! WARNING  sprint {s["sprint"]} is over capacity: {s["planned"]} h planned '
              f'against {s["capacity"]} h available')

    wb = Workbook()

    # ------------------------------------------------------- 1. Read Me
    ws = wb.active
    ws.title = GLOSSARY_SHEET
    rm = content['Read Me']
    sheet_header(ws, rm.get('Title', 'Trial Run Plan'),
                 f'Generated from the plan of record on {ISSUE_DATE}. '
                 f'Never edit this workbook — edit the plan and rebuild.', tab='1F3864')
    widths(ws, [('Item', 26), ('Detail', 150)])
    order = ['Purpose', 'Scope', 'Headline', 'Reading order', 'One caution']
    rows = [[k, rm[k]] for k in order if k in rm]
    r = table(ws, 4, [('', 26), ('', 150)], rows) + 1
    r = section(ws, r, 'Glossary — the only place shorthand appears in this workbook')
    r = table(ws, r, [('Short form', 26), ('Written out', 150)],
              [[a, b] for a, b in content['Glossary']], centre_cols=(0,))
    ws.freeze_panes = 'A4'

    # ------------------------------------------------------- 2. Effort Summary
    ws = wb.create_sheet('Effort Summary')
    sheet_header(ws, 'Effort Summary',
                 f'Where the {TOTAL_HOURS} hours of development effort sit, by block of work '
                 f'and by discipline. Development only — testing, business analysis and '
                 f'contingency are separate and are not included.', tab='2E7D32')
    cols = [('Block of work', 74), ('Hours', 12), ('Days (6.5 h)', 13), ('Share', 10)]
    widths(ws, cols)
    rows = [[b['block'], b['hours'], days(b['hours']), b['share']] for b in plan['blocks']]
    rows.append(['Total', TOTAL_HOURS, days(TOTAL_HOURS), '100 %'])
    r = table(ws, 4, cols, rows, centre_cols=(1, 2, 3),
              total_rows=(len(rows) - 1,)) + 1
    r = note(ws, r, ['The platform is the largest single block and has not been started. '
                     'Reducing screen scope is the weakest lever available to this plan; '
                     'shortening the platform is the strongest.']) + 1
    r = section(ws, r, 'By discipline')
    dcols = [('Discipline', 74), ('Hours', 12), ('Days (6.5 h)', 13), ('Share', 10)]
    drows = [[DISCIPLINE.get(d['code'], d['name']), d['hours'], days(d['hours']), d['share']]
             for d in plan['disciplines']]
    drows.append(['Total', TOTAL_HOURS, days(TOTAL_HOURS), '100 %'])
    table(ws, r, dcols, drows, centre_cols=(1, 2, 3), total_rows=(len(drows) - 1,))
    ws.freeze_panes = 'A5'

    # ------------------------------------------------------- 3. Effort by Phase and Discipline
    ws = wb.create_sheet('Effort by Phase and Discipline')
    sheet_header(ws, 'Effort by Phase and Discipline',
                 'The staffing grid. A block total does not say whether the work is screen '
                 'work or service work; this does. Every row is the sum of its own cells and '
                 'every column sums to its total.', tab='1565C0')
    cols = [('Phase', 46)] + [(DISCIPLINE[c], 22) for c in ('FE', 'BE', 'DB', 'RT')] \
           + [('Total hours', 13), ('Days (6.5 h)', 13), ('Sprint', 26)]
    widths(ws, cols)
    rows = []
    for row in plan['grid']:
        label = phase_label(row['phase'])
        rows.append([label or 'Total'] + [row[k] or None for k in ('fe', 'be', 'db', 'rt')]
                    + [row['total'], days(row['total']), row['sprint']])
    rows[-1][0] = 'Total'
    r = table(ws, 4, cols, rows, centre_cols=(1, 2, 3, 4, 5, 6, 7),
              total_rows=(len(rows) - 1,)) + 1
    note(ws, r, [
        'The platform is three unequal problems that genuinely run in parallel, which is why '
        'an extra person pays for itself in the first sprint and nowhere else.',
        'Real-time and controller integration is concentrated in the platform: after it, only '
        '39 hours remain across five blocks, so that owner picks up database or service work.',
        'Spool completion is the only block that is service-heavy, and it is also last and '
        'gates the finishing line — so that owner holds the end of the schedule.'])
    ws.freeze_panes = 'B5'

    # ------------------------------------------------------- 4. Platform Detail
    ws = wb.create_sheet('Platform Detail')
    sheet_header(ws, 'Platform Detail — Phase 1',
                 f'{PLATFORM_HOURS} hours, 54 per cent of the trial, not yet started. Broken '
                 f'out to deliverable level because a single figure for the platform is not '
                 f'something anyone can act on.', tab='6A1B9A')
    cols = [('Sub-phase', 42), ('Hand-coded basis', 16), ('Faster basis, full scope', 18),
            ('Trial scope', 13), ('Reduction', 14), ('Sprint 1', 11), ('Sprint 2', 11)]
    widths(ws, cols)
    rows = []
    for b in plan['bases']:
        label = phase_label(b['subphase']) or 'Total'
        rows.append([label, b['handcoded'], b['ai'], b['trial'], b['reduction'],
                     b['t1'] or None, b['t2'] or None])
    rows[-1][0] = 'Total'
    r = table(ws, 4, cols, rows, centre_cols=(1, 2, 3, 4, 5, 6),
              total_rows=(len(rows) - 1,)) + 1
    r = note(ws, r, [
        'Three bases, and they must not be mixed. The hand-coded basis is the estimate of '
        'record and includes testing and contingency. The other two are development only.',
        'Only the backend foundation crosses a sprint boundary. The Angular and database '
        'foundations must finish in the first sprint or the whole chain after them slips.']) + 1

    r = section(ws, r, 'All-in, on the trial basis — comparable with the hand-coded column above')
    acols = [('Sub-phase', 42), ('Development', 14), ('Testing', 12), ('Contingency', 13),
             ('All-in hours', 13)]
    arows = []
    for a in plan['allin']:
        label = phase_label(a['subphase'])
        arows.append([label, a['dev'], a['qa'], a['cont'], a['total']])
    arows[-1][0] = 'Total'
    r = table(ws, r, acols, arows, centre_cols=(1, 2, 3, 4),
              total_rows=(len(arows) - 1,)) + 1
    r = note(ws, r, [
        'Quote the development figure against a team roster and the all-in figure against a '
        'budget. Comparing development-only against the hand-coded column overstates the '
        'saving: the like-for-like comparison is 584 against 1,027.']) + 1

    for code in ('1A', '1B', '1C'):
        r = section(ws, r, SUBPHASE[code])
        pcols = [('Deliverable', 62), ('Story', 12), ('Hand-coded', 13), ('Trial hours', 12)]
        prows = [[i['deliverable'], i['story'], i['handcoded'] or None, i['trial']]
                 for i in plan['platform'][code]]
        subtotal = sum(i['trial'] for i in plan['platform'][code])
        prows.append(['Total', '', None, subtotal])
        r = table(ws, r, pcols, prows, centre_cols=(1, 2, 3),
                  total_rows=(len(prows) - 1,)) + 1
    ws.freeze_panes = 'A5'

    # ------------------------------------------------------- 5. Work Items
    ws = wb.create_sheet('Work Items')
    sheet_header(ws, 'Work Items',
                 f'All {len(plan["items"])} pieces of work in the trial, with what each one '
                 f'delivers. Use the filters on the header row to narrow to a sprint or a '
                 f'phase.', tab='7B1FA2')
    cols = [('Story', 12), ('Work item', 40), ('What it delivers', 96),
            ('Technical name', 54), ('Phase or group', 34), ('Sprint', 22),
            ('Hours', 10), ('Days (6.5 h)', 13)]
    widths(ws, cols)
    order = {'T1': 0, 'T2': 1, 'T3': 2}
    items = sorted(plan['items'], key=lambda i: (order[i['sprint']], i['group'], i['story']))
    rows = []
    for i in items:
        e = content['Work Items'][i['story']]
        rows.append([i['story'], e['Work item'], e['Delivers'], titles[i['story']],
                     i['group'], SPRINT_NAME[i['sprint']], i['hours'], days(i['hours'])])
    rows.append(['', 'Total', '', '', '', '', TOTAL_HOURS, days(TOTAL_HOURS)])
    last = table(ws, 4, cols, rows, centre_cols=(0, 5, 6, 7), total_rows=(len(rows) - 1,))
    ws.auto_filter.ref = f'A4:{get_column_letter(len(cols))}{last - 2}'
    ws.freeze_panes = 'C5'

    # ------------------------------------------------------- 6. Sprint Plan
    ws = wb.create_sheet('Sprint Plan')
    sheet_header(ws, 'Sprint Plan',
                 'Four blocks of work from 31 August. The first is the platform; the last is testing '
                 'and sign-off, which cannot share a sprint with development work at any team '
                 'size.', tab='EF6C00')
    cols = [('Sprint', 24), ('Dates', 24), ('Working days', 13), ('Team', 10),
            ('Capacity hours', 14), ('Planned hours', 13), ('Utilisation', 12),
            ('Content', 62)]
    widths(ws, cols)
    rows = [[SPRINT_NAME.get(clean(s['sprint']).strip('*'), s['sprint']), s['dates'],
             s['workdays'], s['team'], s['capacity'], s['planned'], s['util'], s['content']]
            for s in plan['sprints']]
    r = table(ws, 4, cols, rows, centre_cols=(2, 3, 4, 5, 6)) + 1
    r = note(ws, r, ['Development completes at the end of the third sprint. The fourth is '
                     'regression, replacing the assumptions the build stood on, defect '
                     'fixing, and user acceptance testing with sign-off.']) + 1

    r = section(ws, r, 'What each staffing option lands')
    ocols = [('Option', 34), ('Development complete', 22), ('Sign-off', 20), ('Note', 96)]
    r = table(ws, r, ocols, [[o['option'], o['complete'], o['signoff'], o['note']]
                             for o in plan['options']], centre_cols=(1, 2)) + 1

    r = section(ws, r, 'Allowances held outside the hours above')
    rcols = [('Allowance', 46), ('Hours', 14), ('Basis', 96)]
    table(ws, r, rcols, [[x['reserve'], x['hours'], x['basis']] for x in plan['reserves']],
          centre_cols=(1,))
    ws.freeze_panes = 'A5'

    # ------------------------------------------------------- 7. Sprint Allocation
    ws = wb.create_sheet('Sprint Allocation')
    sheet_header(ws, 'Sprint Allocation',
                 'The same work grouped the way it is committed, sprint by sprint, with the '
                 'subtotal for each group. Every sprint reconciles to the Sprint Plan.',
                 tab='00838F')
    cols = [('Group', 40), ('Story', 12), ('Work item', 44), ('Hours', 10), ('Days (6.5 h)', 13)]
    widths(ws, cols)
    r = 4
    for code in ('T1', 'T2', 'T3'):
        total = plan['sprint_totals'][code]
        r = section(ws, r, f'{SPRINT_NAME[code]} — {total} hours')
        rows, groups = [], {}
        for i in [x for x in plan['items'] if x['sprint'] == code]:
            groups.setdefault(i['group'], []).append(i)
        for group in groups:
            for i in sorted(groups[group], key=lambda x: x['story']):
                rows.append([group, i['story'], content['Work Items'][i['story']]['Work item'],
                             i['hours'], days(i['hours'])])
            rows.append([f'{group} — subtotal', '', '',
                         sum(x['hours'] for x in groups[group]),
                         days(sum(x['hours'] for x in groups[group]))])
        rows.append([f'{SPRINT_NAME[code]} — total', '', '', total, days(total)])
        totals = tuple(n for n, v in enumerate(rows) if 'subtotal' in str(v[0])
                       or 'total' in str(v[0]))
        r = table(ws, r, cols, rows, centre_cols=(1, 3, 4), total_rows=totals) + 1
    ws.freeze_panes = 'A5'

    # ------------------------------------------------------- 8. Blockers
    ws = wb.create_sheet('Blockers')
    sheet_header(ws, 'Blockers',
                 'What must close before the work it gates can proceed, and when. Two of the '
                 'five are values or confirmations owed to us from outside; three are ours '
                 'to decide.', tab='C62828')
    cols = [('Number', 10), ('Blocker', 32), ('What we need', 88), ('What it blocks', 44),
            ('Needed by', 22), ('Consequence if late', 70)]
    widths(ws, cols)
    rows = []
    for b in plan['blockers']:
        key = f'B{clean(b["n"])}'
        e = content['Blockers'].get(key, {})
        rows.append([clean(b['n']), e.get('Blocker', ''), e.get('What we need', ''),
                     b['blocks'], b['needed'], e.get('Consequence if late', '')])
    r = table(ws, 4, cols, rows, centre_cols=(0, 4), focus_cols=(4,)) + 1
    # Keyed TIER2, not B5: the plan's blocker table gained a real row 5 (G38) on
    # 15 Aug 2026, and the second tier had been squatting on that key.
    tier2 = content['Blockers'].get('TIER2', {})
    if tier2:
        r = section(ws, r, 'Second tier — none of these stops the build')
        r = table(ws, r, [('Blocker', 32), ('What we need', 88),
                          ('Consequence if late', 70)],
                  [[tier2.get('Blocker', ''), tier2.get('What we need', ''),
                    tier2.get('Consequence if late', '')]])
    ws.freeze_panes = 'B5'

    # ------------------------------------------------------- 9. Deferred Items
    ws = wb.create_sheet('Deferred Items')
    sheet_header(ws, 'Deferred Items',
                 f'{DEFERRED_HOURS} hours deliberately outside the trial and why each is safe '
                 f'to defer. All of it remains in the wider delivery — none of it is cancelled.',
                 tab='546E7A')
    cols = [('Deferred', 56), ('Hours', 10), ('Days (6.5 h)', 13),
            ('Why it is safe to defer for a trial', 128)]
    widths(ws, cols)
    rows = [[d['item'], d['hours'], days(d['hours']), d['reason']] for d in plan['deferred']]
    rows.append(['Total', DEFERRED_HOURS, days(DEFERRED_HOURS), ''])
    last = table(ws, 4, cols, rows, centre_cols=(1, 2), total_rows=(len(rows) - 1,))
    ws.auto_filter.ref = f'A4:D{last - 2}'
    ws.freeze_panes = 'A5'

    # ------------------------------------------------------- 10. Removal Impact
    ws = wb.create_sheet('Removal Impact')
    sheet_header(ws, 'Removal Impact',
                 'Two screens were removed from the trial on 14 August. Neither was a clean '
                 'subtraction. These are the consequences, stated so they are decisions rather '
                 'than discoveries at sign-off.', tab='AD1457')
    cols = [('Removed', 40), ('Consequence', 92), ('How it is handled', 108)]
    widths(ws, cols)
    rows = [[x['removal'], x['consequence'], x['handling']] for x in plan['removals']]
    last = table(ws, 4, cols, rows)
    r = note(ws, last, [
        'The two consequences to carry into a sign-off conversation: the trial exercises no '
        'weld traceability, so the genealogy behind the welding-wire certificates is not '
        'demonstrated; and no screen in the trial displays an alert, so a rejection notifying '
        'a supervisor is verifiable only in the message log.'])
    ws.auto_filter.ref = f'A4:C{last - 1}'
    ws.freeze_panes = 'B5'

    wb.save(out_path)

    hits = check_abbreviations(out_path)
    if hits:
        print('FATAL - build refused, abbreviations reached the workbook:')
        for h in hits:
            print(f'  * {h}')
        os.remove(out_path)
        sys.exit(1)

    return dict(sheets=[w.title for w in wb.worksheets], items=len(plan['items']),
                hours=TOTAL_HOURS, deferred=DEFERRED_HOURS,
                blockers=len(plan['blockers']))


# Re-baselined 25 Aug 2026 with TrialRunPlan.md section 2.1. These strings are NOT read from the
# plan - they are the row labels on the Sprint Allocation and Work Items sheets, so they have to
# be moved by hand whenever the calendar in 2.1 moves.
SPRINT_NAME = {'T1': 'Sprint 1 (31 August – 29 September)',
               'T2': 'Sprint 2 (30 September – 20 October)',
               'T3': 'Sprint 3 (21 October – 3 November)',
               'T4': 'Sprint 4 (4–16 November)'}


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT
    stats = build(out)
    print(f'wrote {os.path.relpath(out, ROOT)}')
    print(f'  work items           : {stats["items"]}')
    print(f'  development effort   : {stats["hours"]} hours '
          f'({days(stats["hours"], 1)} days)')
    print(f'  deferred             : {stats["deferred"]} hours')
    print(f'  blockers             : {stats["blockers"]}')
    print(f'  sheets               : {" · ".join(stats["sheets"])}')
    print(f'  abbreviation scan    : clean')


if __name__ == '__main__':
    main()
